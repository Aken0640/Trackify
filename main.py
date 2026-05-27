"""
Trackify v3.1 — Complete Personal Finance Manager
Fixes: Instant theme switching (full UI rebuild), analytics month arrow nav,
uniform calendar cells (total inc/exp only), editable goals, dashboard grouped
by day with month arrow nav (no date pickers / no min-max), Ctrl+F badge on
search, Ctrl+N tooltip on FAB, library install note on export page.
"""

import tkinter as tk
from tkinter import ttk, messagebox
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import csv, os, json, calendar, shutil
from tkcalendar import DateEntry
from datetime import datetime, date, timedelta
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    HAS_PDF = True
except ImportError:
    HAS_PDF = False


# ─────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────

def fmt(amount: float) -> str:
    return f"₹{amount:,.2f}"

def center_window(win, w, h, relative_to=None):
    if relative_to:
        x = relative_to.winfo_x() + relative_to.winfo_width()//2 - w//2
        y = relative_to.winfo_y() + relative_to.winfo_height()//2 - h//2
    else:
        sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
        x, y = (sw-w)//2, (sh-h)//2
    win.geometry(f"{w}x{h}+{x}+{y}")


# ─────────────────────────────────────────────
#  Theme
# ─────────────────────────────────────────────

THEMES = {
    "Dark": {
        "BG":"#0D0D0D","PANEL":"#161616","CARD":"#1E1E1E","BORDER":"#2A2A2A",
        "SIDEBAR":"#111111","SIDEBAR_ACTIVE":"#1E1E1E",
        "ACCENT":"#00FFFF","GREEN":"#00CC88","RED":"#FF4D4D",
        "ORANGE":"#FF944D","TEXT":"#FFFFFF","SUBTEXT":"#888888",
        "CHART":["#FF4D4D","#FF944D","#4DA6FF","#CC66FF","#00CCCC","#FF66B2","#00CC88","#FFD700"],
    },
    "Light": {
        "BG":"#EEF2F7","PANEL":"#FFFFFF","CARD":"#FFFFFF","BORDER":"#CBD5E1",
        "SIDEBAR":"#1E293B","SIDEBAR_ACTIVE":"#334155",
        "ACCENT":"#1D4ED8","GREEN":"#15803D","RED":"#B91C1C",
        "ORANGE":"#C2410C","TEXT":"#0F172A","SUBTEXT":"#475569",
        "CHART":["#B91C1C","#C2410C","#1D4ED8","#6D28D9","#0369A1","#BE185D","#15803D","#B45309"],
    },
    "Neon": {
        "BG":"#050510","PANEL":"#0A0A1A","CARD":"#0F0F20","BORDER":"#1A1A3A",
        "SIDEBAR":"#070712","SIDEBAR_ACTIVE":"#0F0F25",
        "ACCENT":"#FF00FF","GREEN":"#00FF88","RED":"#FF0055",
        "ORANGE":"#FF8800","TEXT":"#E0E0FF","SUBTEXT":"#6666AA",
        "CHART":["#FF00FF","#FF0055","#0088FF","#FF8800","#00FFFF","#FF44AA","#00FF88","#FFFF00"],
    },
    "Cyberpunk": {
        "BG":"#0A0A0F","PANEL":"#12121A","CARD":"#1A1A25","BORDER":"#2A2A3A",
        "SIDEBAR":"#0D0D15","SIDEBAR_ACTIVE":"#1A1A25",
        "ACCENT":"#F7FF00","GREEN":"#00FF9F","RED":"#FF2D55",
        "ORANGE":"#FF6B00","TEXT":"#E8E8F0","SUBTEXT":"#7070A0",
        "CHART":["#F7FF00","#FF2D55","#00BFFF","#BF00FF","#00FF9F","#FF6B00","#FF00BF","#00FFFF"],
    },
}

class Theme:
    _name = "Dark"
    _t    = THEMES["Dark"]
    @classmethod
    def set(cls, name):
        cls._name = name
        cls._t    = THEMES.get(name, THEMES["Dark"])
    @classmethod
    def name(cls): return cls._name
    @classmethod
    def get(cls, key): return cls._t.get(key, "#000000")


# ─────────────────────────────────────────────
#  Tooltip
# ─────────────────────────────────────────────

class Tooltip:
    def __init__(self, widget, text):
        self.widget, self.text, self.tip = widget, text, None
        widget.bind("<Enter>", self._show)
        widget.bind("<Leave>", self._hide)
    def _show(self, _=None):
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 4
        self.tip = tk.Toplevel(self.widget)
        self.tip.wm_overrideredirect(True)
        self.tip.wm_geometry(f"+{x}+{y}")
        tk.Label(self.tip, text=self.text, justify="left",
                 bg="#1A1A1A", fg="#FFFFFF", font=("Consolas", 9),
                 relief="flat", padx=10, pady=6).pack()
    def _hide(self, _=None):
        if self.tip: self.tip.destroy(); self.tip = None


# ─────────────────────────────────────────────
#  Config
# ─────────────────────────────────────────────

CONFIG_FILE = "trackify_config.json"

def load_config():
    d = {"theme":"Dark","geometry":"1100x700","budgets":{},"goals":[],"recurring":[]}
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE) as f: d.update(json.load(f))
        except Exception: pass
    return d

def save_config(cfg):
    with open(CONFIG_FILE,"w") as f: json.dump(cfg, f, indent=2)


# ─────────────────────────────────────────────
#  Intro
# ─────────────────────────────────────────────

class IntroPage:
    def __init__(self, root):
        self.root = root
        root.title("Trackify"); root.configure(bg=Theme.get("BG")); root.resizable(False,False)
        center_window(root,480,540)
        self.frame = tk.Frame(root, bg=Theme.get("BG")); self.frame.pack(expand=True)

        self.title_lbl = tk.Label(self.frame, text="Trackify", font=("Helvetica",68,"bold"),
                                  fg=Theme.get("ACCENT"), bg=Theme.get("BG"))
        self.title_lbl.pack(pady=(60,4))
        tk.Frame(self.frame, bg=Theme.get("ACCENT"), height=2, width=280).pack()
        tk.Label(self.frame, text="The Complete Finance Manager", font=("Helvetica",14),
                 fg=Theme.get("SUBTEXT"), bg=Theme.get("BG")).pack(pady=(10,6))
        tk.Label(self.frame, text="v3.1 · Budget · Goals · Analytics · Export",
                 font=("Helvetica",9), fg=Theme.get("SUBTEXT"), bg=Theme.get("BG")).pack(pady=(0,30))

        btn = tk.Button(self.frame, text="  Get Started  →  ", font=("Helvetica",13,"bold"),
                        bg=Theme.get("ACCENT"), fg=Theme.get("BG"), relief="flat",
                        padx=24, pady=12, cursor="hand2", command=self.start_app)
        btn.pack()
        btn.bind("<Enter>", lambda _: btn.config(bg="#66FFFF"))
        btn.bind("<Leave>", lambda _: btn.config(bg=Theme.get("ACCENT")))
        Tooltip(btn, "Launch Trackify  [Enter]")

        tk.Label(self.frame, text="v3.1 · Personal Finance Manager",
                 font=("Helvetica",9), fg="#383838", bg=Theme.get("BG")).pack(pady=(22,0))

        self._gc = ["#00FFFF","#33FFFF","#66FFFF","#99FFFF","#66FFFF","#33FFFF"]
        self._gi = 0; self._gj = None; self._glow()

        # Bind Enter key to launch the app
        root.bind("<Return>", lambda _: self.start_app())

    def _glow(self):
        if not self.title_lbl.winfo_exists(): return
        self.title_lbl.config(fg=self._gc[self._gi])
        self._gi = (self._gi+1) % len(self._gc)
        self._gj = self.root.after(380, self._glow)

    def start_app(self):
        if self._gj: self.root.after_cancel(self._gj)
        self.frame.destroy()
        ExpenseTracker(self.root)


# ─────────────────────────────────────────────
#  Main App
# ─────────────────────────────────────────────

class ExpenseTracker:
    FILE = "transactions.csv"
    COLS = ("Date","Type","Category","Amount","Notes")
    INCOME_CATS  = ["Allowance","Salary","Petty Cash","Bonus","Freelance","Other"]
    EXPENSE_CATS = ["Food","Transport","Shopping","Bills","Entertainment","Health","Other"]

    def __init__(self, root):
        self.root = root
        self.cfg  = load_config()
        Theme.set(self.cfg.get("theme","Dark"))

        root.title("Trackify — Expense Tracker")
        root.configure(bg=Theme.get("BG"))
        root.resizable(True,True)
        root.geometry(self.cfg.get("geometry","1100x700"))
        center_window(root,1100,700)

        self.balance=0.0; self.total_income=0.0; self.total_expense=0.0
        self.expenses={}; self.all_transactions=[]
        self._undo_buffer=None; self._undo_job=None

        self.budgets   = self.cfg.get("budgets",{})
        self.goals     = self.cfg.get("goals",[])
        self.recurring = self.cfg.get("recurring",[])

        # Dashboard month view state
        self._dash_year  = date.today().year
        self._dash_month = date.today().month

        self._ensure_csv()
        self._build_all()
        self._load_transactions()
        self._process_recurring()
        self._bind_shortcuts()
        root.protocol("WM_DELETE_WINDOW", self._on_close)

    # ── internal rebuild (called on theme switch) ──────────────────────
    def _full_rebuild(self):
        """Destroy all widgets and rebuild with new theme colours."""
        for w in self.root.winfo_children():
            w.destroy()
        self.balance=0.0; self.total_income=0.0; self.total_expense=0.0
        self.expenses={}; self.all_transactions=[]
        self._undo_buffer=None; self._undo_job=None
        self._dash_year  = date.today().year
        self._dash_month = date.today().month
        self._build_all()
        self._load_transactions()
        self._bind_shortcuts()

    def _build_all(self):
        self.root.configure(bg=Theme.get("BG"))
        self._build_styles()
        self._nav_buttons  = {}
        self._current_page = None
        self._pages        = {}
        self._page_builders = {
            "dashboard": self._build_dashboard_page,
            "analytics": self._build_analytics_page,
            "budget":    self._build_budget_page,
            "calendar":  self._build_calendar_page,
            "goals":     self._build_goals_page,
            "recurring": self._build_recurring_page,
            "export":    self._build_export_page,
            "settings":  self._build_settings_page,
        }
        self._main = tk.Frame(self.root, bg=Theme.get("BG"))
        self._main.pack(fill="both", expand=True)
        self._build_sidebar()
        self._content_host = tk.Frame(self._main, bg=Theme.get("BG"))
        self._content_host.pack(side="left", fill="both", expand=True)
        self._show_page("dashboard")

    def _ensure_csv(self):
        if not os.path.exists(self.FILE):
            with open(self.FILE,"w",newline="") as f:
                csv.writer(f).writerow(self.COLS)

    def _build_styles(self):
        s = ttk.Style(); s.theme_use("default")
        bg,card,text,accent,subtext,panel = (
            Theme.get("BG"),Theme.get("CARD"),Theme.get("TEXT"),
            Theme.get("ACCENT"),Theme.get("SUBTEXT"),Theme.get("PANEL"))
        s.configure("Treeview", background=card, fieldbackground=card,
                    foreground=text, rowheight=28, font=("Consolas",10))
        s.configure("Treeview.Heading", background=panel, foreground=accent,
                    font=("Arial",10,"bold"), relief="flat")
        s.map("Treeview",
              background=[("selected",Theme.get("BORDER"))],
              foreground=[("selected",accent)])
        s.configure("TCombobox", fieldbackground=card, background=card,
                    foreground=text, selectbackground=card, selectforeground=text,
                    arrowcolor=accent, borderwidth=0, relief="flat")
        s.map("TCombobox",
              fieldbackground=[("readonly",card),("disabled",panel)],
              foreground=[("readonly",text),("disabled",subtext)],
              selectbackground=[("readonly",card)],
              selectforeground=[("readonly",text)])
        self.root.option_add("*TCombobox*Listbox.background", card)
        self.root.option_add("*TCombobox*Listbox.foreground", text)
        self.root.option_add("*TCombobox*Listbox.selectBackground", accent)
        self.root.option_add("*TCombobox*Listbox.selectForeground", bg)
        s.configure("Vertical.TScrollbar", background=card,
                    troughcolor=bg, borderwidth=0, arrowcolor=accent)

    def _bind_shortcuts(self):
        self.root.bind("<Control-n>", lambda _: self._open_add_window())
        self.root.bind("<Control-N>", lambda _: self._open_add_window())
        self.root.bind("<Control-f>", lambda _: self._focus_search())
        self.root.bind("<Control-F>", lambda _: self._focus_search())
        self.root.bind("<Delete>",    lambda _: self._delete_selected())
        self.root.bind("<Control-z>", lambda _: self._undo_delete())
        self.root.bind("<Control-Z>", lambda _: self._undo_delete())
        self.root.bind("<Control-e>", lambda _: self._show_page("export"))
        self.root.bind("<F5>",        lambda _: self._refresh_dash_table())

    def _focus_search(self):
        if hasattr(self,"_search_entry"):
            self._search_entry.focus_set()

    # ─── Sidebar ──────────────────────────────────────────────────────

    def _build_sidebar(self):
        sb = tk.Frame(self._main, bg=Theme.get("SIDEBAR"), width=190)
        sb.pack(side="left", fill="y"); sb.pack_propagate(False)

        tk.Label(sb, text="⚡ Trackify", font=("Arial",15,"bold"),
                 fg=Theme.get("ACCENT"), bg=Theme.get("SIDEBAR"), pady=18).pack(fill="x",padx=16)
        tk.Frame(sb, bg=Theme.get("BORDER"), height=1).pack(fill="x",padx=10,pady=(0,8))

        nav_items = [
            ("dashboard","🏠  Dashboard"),("analytics","📊  Analytics"),
            ("budget","💰  Budget"),("calendar","📅  Calendar"),
            ("goals","🎯  Goals"),("recurring","🔄  Recurring"),
            ("export","📤  Export"),("settings","⚙️  Settings"),
        ]
        for key,label in nav_items:
            btn = tk.Button(sb, text=label, font=("Arial",10),
                            bg=Theme.get("SIDEBAR"), fg=Theme.get("TEXT"),
                            relief="flat", anchor="w", padx=16, pady=10,
                            cursor="hand2", command=lambda k=key: self._show_page(k),
                            activebackground=Theme.get("SIDEBAR_ACTIVE"),
                            activeforeground=Theme.get("ACCENT"))
            btn.pack(fill="x")
            btn.bind("<Enter>", lambda e,b=btn: b.config(bg=Theme.get("SIDEBAR_ACTIVE")))
            btn.bind("<Leave>", lambda e,b=btn,k=key: b.config(
                bg=Theme.get("SIDEBAR_ACTIVE") if self._current_page==k else Theme.get("SIDEBAR")))
            self._nav_buttons[key] = btn

        # Theme selector at bottom — custom upward-opening dropdown
        tk.Frame(sb, bg=Theme.get("BORDER"), height=1).pack(fill="x",padx=10,pady=8,side="bottom")
        tf = tk.Frame(sb, bg=Theme.get("SIDEBAR"))
        tf.pack(side="bottom", fill="x", padx=10, pady=4)
        tk.Label(tf, text="Theme", font=("Arial",8),
                 fg=Theme.get("SUBTEXT"), bg=Theme.get("SIDEBAR")).pack(anchor="w")

        self._theme_var = tk.StringVar(value=Theme.name())

        def _open_theme_menu():
            btn_widget = _theme_btn
            x = btn_widget.winfo_rootx()
            # Calculate y position ABOVE the button
            popup_h = len(THEMES) * 32 + 4
            y = btn_widget.winfo_rooty() - popup_h - 2

            menu_win = tk.Toplevel(self.root)
            menu_win.wm_overrideredirect(True)
            menu_win.wm_geometry(f"+{x}+{y}")
            menu_win.configure(bg=Theme.get("BORDER"))

            def _pick(name):
                self._theme_var.set(name)
                menu_win.destroy()
                if name != Theme.name():
                    Theme.set(name)
                    self.cfg["theme"] = name
                    save_config(self.cfg)
                    self._full_rebuild()

            for name in THEMES.keys():
                is_active = (name == Theme.name())
                item = tk.Button(menu_win, text=name,
                                 font=("Arial",10,"bold" if is_active else "normal"),
                                 bg=Theme.get("SIDEBAR_ACTIVE") if is_active else Theme.get("PANEL"),
                                 fg=Theme.get("ACCENT") if is_active else Theme.get("TEXT"),
                                 relief="flat", anchor="w", padx=14, pady=6,
                                 cursor="hand2", command=lambda n=name: _pick(n))
                item.pack(fill="x", padx=1, pady=1)
                item.bind("<Enter>", lambda e, b=item: b.config(bg=Theme.get("SIDEBAR_ACTIVE")))
                item.bind("<Leave>", lambda e, b=item, n=name: b.config(
                    bg=Theme.get("SIDEBAR_ACTIVE") if n==Theme.name() else Theme.get("PANEL")))

            menu_win.bind("<FocusOut>", lambda e: menu_win.destroy())
            menu_win.focus_set()

        _theme_btn = tk.Button(tf, textvariable=self._theme_var,
                               font=("Arial",9,"bold"),
                               bg=Theme.get("CARD"), fg=Theme.get("ACCENT"),
                               relief="flat", anchor="w", padx=10, pady=6,
                               cursor="hand2", command=_open_theme_menu)
        _theme_btn.pack(fill="x", pady=2)
        _theme_btn.bind("<Enter>", lambda e: _theme_btn.config(bg=Theme.get("SIDEBAR_ACTIVE")))
        _theme_btn.bind("<Leave>", lambda e: _theme_btn.config(bg=Theme.get("CARD")))
        self._sidebar = sb

    def _apply_theme_instant(self):
        name = self._theme_var.get()
        if name == Theme.name(): return
        Theme.set(name)
        self.cfg["theme"] = name
        save_config(self.cfg)
        self._full_rebuild()          # ← instant full redraw
    def _show_page(self, key):
        if self._current_page and self._current_page in self._nav_buttons:
            self._nav_buttons[self._current_page].config(bg=Theme.get("SIDEBAR"),
                                                          fg=Theme.get("TEXT"))
        self._current_page = key
        self._nav_buttons[key].config(bg=Theme.get("SIDEBAR_ACTIVE"),
                                       fg=Theme.get("ACCENT"))
        for pg in self._pages.values(): pg.pack_forget()
        if key not in self._pages:
            frame = tk.Frame(self._content_host, bg=Theme.get("BG"))
            self._pages[key] = frame
            self._page_builders[key](frame)
        self._pages[key].pack(fill="both", expand=True)
        if key == "analytics":  self._refresh_analytics()
        elif key == "budget":   self._refresh_budget_page()
        elif key == "calendar": self._refresh_calendar()
        elif key == "goals":    self._refresh_goals_page()
        elif key == "dashboard":self._refresh_dash_table()

    # ─── Dashboard ────────────────────────────────────────────────────

    def _build_dashboard_page(self, frame):
        # ── top bar with month nav ──
        bar = tk.Frame(frame, bg=Theme.get("PANEL"), pady=10)
        bar.pack(fill="x")
        tk.Label(bar, text="Dashboard", font=("Arial",14,"bold"),
                 fg=Theme.get("ACCENT"), bg=Theme.get("PANEL")).pack(side="left",padx=20)

        # month nav (right side of topbar)
        nav = tk.Frame(bar, bg=Theme.get("PANEL"))
        nav.pack(side="right", padx=20)
        tk.Button(nav, text="◀", font=("Arial",11), bg=Theme.get("CARD"),
                  fg=Theme.get("TEXT"), relief="flat", cursor="hand2",
                  command=lambda: self._dash_nav(-1)).pack(side="left", padx=2)
        self._dash_month_lbl = tk.Label(nav, text="", font=("Arial",11,"bold"),
                                        fg=Theme.get("TEXT"), bg=Theme.get("PANEL"), width=16)
        self._dash_month_lbl.pack(side="left")
        tk.Button(nav, text="▶", font=("Arial",11), bg=Theme.get("CARD"),
                  fg=Theme.get("TEXT"), relief="flat", cursor="hand2",
                  command=lambda: self._dash_nav(1)).pack(side="left", padx=2)

        # ── summary cards ──
        cards_row = tk.Frame(frame, bg=Theme.get("BG"))
        cards_row.pack(fill="x", padx=16, pady=(12,4))
        self._bal_lbl = self._card(cards_row,"💰  Balance", fmt(0), Theme.get("ACCENT"))
        self._inc_lbl = self._card(cards_row,"📈  Income",  fmt(0), Theme.get("GREEN"))
        self._exp_lbl = self._card(cards_row,"📉  Expenses",fmt(0), Theme.get("RED"))
        self._sav_lbl = self._card(cards_row,"💾  Savings %","0.0%",Theme.get("ORANGE"))
        for w in cards_row.winfo_children():
            w.pack(side="left", expand=True, fill="both", padx=4)

        # ── toolbar (filter + search only, no date pickers / no min-max) ──
        toolbar = tk.Frame(frame, bg=Theme.get("BG"))
        toolbar.pack(fill="x", padx=16, pady=(6,2))

        tk.Label(toolbar, text="Filter:", font=("Arial",9),
                 fg=Theme.get("SUBTEXT"), bg=Theme.get("BG")).pack(side="left")
        self._filter = tk.StringVar(value="All")
        fcb = ttk.Combobox(toolbar, textvariable=self._filter,
                           values=["All","Income","Expense"], state="readonly", width=9)
        fcb.pack(side="left", padx=(4,14))
        self._filter.trace_add("write", lambda *_: self._refresh_dash_table())

        # ── Search box with inline Ctrl+F hint ──
        tk.Label(toolbar, text="🔍", font=("Arial",10),
                 fg=Theme.get("SUBTEXT"), bg=Theme.get("BG")).pack(side="left")

        # Container so we can overlay the hint label inside the entry
        search_wrap = tk.Frame(toolbar, bg=Theme.get("CARD"))
        search_wrap.pack(side="left", padx=(4,14), ipady=2)

        self._search = tk.StringVar()
        self._search.trace_add("write", lambda *_: self._refresh_dash_table())
        self._search_entry = tk.Entry(search_wrap, textvariable=self._search,
                                      bg=Theme.get("CARD"), fg=Theme.get("TEXT"),
                                      insertbackground=Theme.get("TEXT"),
                                      relief="flat", font=("Arial",10), width=22,
                                      bd=0)
        self._search_entry.pack(side="left", ipady=4, padx=(6,0))
        Tooltip(self._search_entry, "Search transactions  [Ctrl+F]")

        # Ctrl+F hint inside the box — shown when empty, hidden while typing
        _hint = tk.Label(search_wrap, text="Ctrl+F", font=("Consolas",8),
                         fg=Theme.get("ACCENT"), bg=Theme.get("CARD"), padx=4)
        _hint.pack(side="left", padx=(0,4))

        def _on_search_change(*_):
            self._refresh_dash_table()
            _hint.config(fg=Theme.get("CARD") if self._search.get() else Theme.get("ACCENT"))

        self._search.trace_add("write", lambda *_: _on_search_change())

        # Right-side: only Chart button (Delete/Edit removed — use ✕/✏ on each row)
        b = tk.Button(toolbar, text="📊  Chart", font=("Arial",9,"bold"),
                      bg=Theme.get("CARD"), fg=Theme.get("TEXT"), relief="flat",
                      padx=10, pady=4, cursor="hand2", command=self._show_chart,
                      activebackground=Theme.get("ACCENT"),
                      activeforeground=Theme.get("BG"))
        b.pack(side="right", padx=(3,0))
        Tooltip(b, "Open expense charts")

        # ── undo bar (hidden) ──
        self._undo_bar = tk.Frame(frame, bg=Theme.get("ORANGE"), pady=4)
        self._undo_lbl = tk.Label(self._undo_bar, text="", font=("Arial",9),
                                  fg=Theme.get("BG"), bg=Theme.get("ORANGE"))
        self._undo_lbl.pack(side="left", padx=14)
        ub = tk.Button(self._undo_bar, text="Undo  ↩", font=("Arial",9,"bold"),
                       bg=Theme.get("BG"), fg=Theme.get("ORANGE"),
                       relief="flat", padx=10, pady=2, cursor="hand2",
                       command=self._undo_delete)
        ub.pack(side="right", padx=14)
        Tooltip(ub, "Undo last delete  [Ctrl+Z]")

        # ── grouped-by-day transaction area ──
        outer = tk.Frame(frame, bg=Theme.get("BG"))
        outer.pack(fill="both", expand=True, padx=16, pady=6)

        canvas_wrap = tk.Canvas(outer, bg=Theme.get("BG"), highlightthickness=0)
        vsb = ttk.Scrollbar(outer, orient="vertical", command=canvas_wrap.yview)
        canvas_wrap.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas_wrap.pack(side="left", fill="both", expand=True)

        self._dash_inner = tk.Frame(canvas_wrap, bg=Theme.get("BG"))
        self._dash_canvas_window = canvas_wrap.create_window((0,0), window=self._dash_inner, anchor="nw")

        def _on_frame_configure(e):
            canvas_wrap.configure(scrollregion=canvas_wrap.bbox("all"))
        def _on_canvas_configure(e):
            canvas_wrap.itemconfig(self._dash_canvas_window, width=e.width)

        self._dash_inner.bind("<Configure>", _on_frame_configure)
        canvas_wrap.bind("<Configure>", _on_canvas_configure)

        # Scroll with mouse wheel on canvas AND any child widget inside it
        def _dash_scroll(e):
            canvas_wrap.yview_scroll(-1*(e.delta//120), "units")
        canvas_wrap.bind("<MouseWheel>", _dash_scroll)
        # Re-bind children as they are created in _refresh_dash_table
        self._dash_scroll_fn = _dash_scroll
        self._dash_canvas = canvas_wrap

        # ── hidden treeview (used only for selection state / delete / edit) ──
        self.table = ttk.Treeview(frame, columns=self.COLS, show="headings", height=0)
        for col in self.COLS:
            self.table.heading(col, text=col)
            self.table.column(col, width=100)
        self.table.tag_configure("income",  foreground=Theme.get("GREEN"))
        self.table.tag_configure("expense", foreground=Theme.get("RED"))
        # (not packed — acts as data store only)

        # ── status bar ──
        status_bar = tk.Frame(frame, bg=Theme.get("PANEL"), pady=3)
        status_bar.pack(fill="x", side="bottom")
        self._status = tk.StringVar(value="  Ready")
        tk.Label(status_bar, textvariable=self._status, font=("Arial",9),
                 fg=Theme.get("SUBTEXT"), bg=Theme.get("PANEL"),
                 anchor="w", padx=12).pack(side="left")
        tk.Label(status_bar,
                 text="Ctrl+N: Add  |  Ctrl+F: Search  |  Del: Delete  |  Ctrl+Z: Undo",
                 font=("Consolas",8), fg=Theme.get("SUBTEXT"),
                 bg=Theme.get("PANEL"), anchor="e", padx=12).pack(side="right")

        # ── FAB ──
        fab = tk.Button(frame, text="+", font=("Arial",20,"bold"),
                        bg=Theme.get("ACCENT"), fg=Theme.get("BG"),
                        width=3, relief="flat", cursor="hand2",
                        command=self._open_add_window)
        fab.place(relx=0.96, rely=0.93, anchor="center")
        fab.bind("<Enter>", lambda _: fab.config(bg="#66FFFF"))
        fab.bind("<Leave>", lambda _: fab.config(bg=Theme.get("ACCENT")))

        # Manual tooltip for FAB (Tooltip class misses it when placed over canvas)
        _fab_tip = None
        def _fab_enter(_):
            nonlocal _fab_tip
            _fab_tip = tk.Toplevel(fab)
            _fab_tip.wm_overrideredirect(True)
            x = fab.winfo_rootx() - 110
            y = fab.winfo_rooty() - 34
            _fab_tip.wm_geometry(f"+{x}+{y}")
            tk.Label(_fab_tip, text="Add new transaction  [Ctrl+N]",
                     bg="#1A1A1A", fg="#FFFFFF", font=("Consolas",9),
                     relief="flat", padx=10, pady=6).pack()
        def _fab_leave(_):
            nonlocal _fab_tip
            if _fab_tip: _fab_tip.destroy(); _fab_tip = None
        fab.bind("<Enter>", lambda e: (_fab_enter(e), fab.config(bg="#66FFFF")))
        fab.bind("<Leave>", lambda e: (_fab_leave(e), fab.config(bg=Theme.get("ACCENT"))))

    def _dash_nav(self, delta):
        self._dash_month += delta
        if self._dash_month > 12: self._dash_month=1;  self._dash_year+=1
        elif self._dash_month < 1: self._dash_month=12; self._dash_year-=1
        self._refresh_dash_table()

    def _refresh_dash_table(self):
        if not hasattr(self,"_dash_inner"): return
        y, m = self._dash_year, self._dash_month
        if hasattr(self,"_dash_month_lbl"):
            self._dash_month_lbl.config(text=f"{calendar.month_name[m]} {y}")

        f = self._filter.get() if hasattr(self,"_filter") else "All"
        q = self._search.get().lower() if hasattr(self,"_search") else ""

        # Group matching transactions by day
        daily = defaultdict(list)
        for tx in self.all_transactions:
            try: d = datetime.strptime(tx["Date"],"%d-%m-%Y")
            except: continue
            if d.year!=y or d.month!=m: continue
            if f!="All" and tx["Type"]!=f: continue
            if q and q not in " ".join(tx.values()).lower(): continue
            daily[d.day].append(tx)

        for w in self._dash_inner.winfo_children(): w.destroy()

        # Helper: bind mousewheel on a widget and all its descendants
        def _bind_scroll(widget):
            if hasattr(self, "_dash_scroll_fn"):
                widget.bind("<MouseWheel>", self._dash_scroll_fn)
                for child in widget.winfo_children():
                    _bind_scroll(child)

        if not daily:
            tk.Label(self._dash_inner,
                     text=f"No transactions for {calendar.month_name[m]} {y}",
                     font=("Arial",11), fg=Theme.get("SUBTEXT"),
                     bg=Theme.get("BG")).pack(pady=30)
            return

        for day in sorted(daily.keys(), reverse=True):
            txs  = daily[day]
            day_str = f"{day:02d}-{m:02d}-{y:04d}"
            day_inc = sum(float(t["Amount"]) for t in txs if t["Type"]=="Income")
            day_exp = sum(float(t["Amount"]) for t in txs if t["Type"]=="Expense")

            # Day header
            hdr = tk.Frame(self._dash_inner, bg=Theme.get("PANEL"), pady=4)
            hdr.pack(fill="x", pady=(6,1))
            try:
                wd = datetime.strptime(day_str,"%d-%m-%Y").strftime("%A, %d %B %Y")
            except: wd = day_str
            tk.Label(hdr, text=wd, font=("Arial",9,"bold"),
                     fg=Theme.get("ACCENT"), bg=Theme.get("PANEL"),
                     padx=12).pack(side="left")
            summary_parts = []
            if day_inc: summary_parts.append(f"+{fmt(day_inc)}")
            if day_exp: summary_parts.append(f"-{fmt(day_exp)}")
            tk.Label(hdr, text="  ".join(summary_parts), font=("Arial",9),
                     fg=Theme.get("SUBTEXT"), bg=Theme.get("PANEL"),
                     padx=12).pack(side="right")
            _bind_scroll(hdr)

            # Transaction rows
            for tx in txs:
                color = Theme.get("GREEN") if tx["Type"]=="Income" else Theme.get("RED")
                row = tk.Frame(self._dash_inner, bg=Theme.get("CARD"), pady=6, padx=14)
                row.pack(fill="x", pady=1)

                tk.Label(row, text=f"{'▲' if tx['Type']=='Income' else '▼'}",
                         font=("Arial",10,"bold"), fg=color,
                         bg=Theme.get("CARD"), width=2).pack(side="left")
                tk.Label(row, text=tx["Category"], font=("Arial",10),
                         fg=Theme.get("TEXT"), bg=Theme.get("CARD"),
                         width=14, anchor="w").pack(side="left")
                tk.Label(row, text=fmt(float(tx["Amount"])), font=("Arial",10,"bold"),
                         fg=color, bg=Theme.get("CARD"), width=14).pack(side="left")
                if tx.get("Notes"):
                    tk.Label(row, text=tx["Notes"], font=("Arial",9),
                             fg=Theme.get("SUBTEXT"), bg=Theme.get("CARD")).pack(side="left",padx=8)

                # Edit/delete buttons + double-click to edit
                def _make_edit(t=tx):
                    self._edit_tx_direct(t)
                def _make_del(t=tx):
                    self._delete_tx_direct(t)

                edit_btn = tk.Button(row, text="✏", font=("Arial",9),
                                     bg=Theme.get("CARD"), fg=Theme.get("ACCENT"),
                                     relief="flat", cursor="hand2", command=_make_edit)
                edit_btn.pack(side="right", padx=2)
                Tooltip(edit_btn, "Edit  [double-click]")
                del_btn = tk.Button(row, text="✕", font=("Arial",9),
                                    bg=Theme.get("CARD"), fg=Theme.get("RED"),
                                    relief="flat", cursor="hand2", command=_make_del)
                del_btn.pack(side="right", padx=2)
                Tooltip(del_btn, "Delete this entry")

                # Double-click anywhere on the row card opens edit popup
                for widget in [row] + list(row.winfo_children()):
                    widget.bind("<Double-Button-1>", lambda e, t=tx: self._edit_tx_direct(t))
                _bind_scroll(row)

    def _card(self, parent, title, value, color):
        card = tk.Frame(parent, bg=Theme.get("CARD"), pady=12, padx=16, relief="flat")
        tk.Label(card, text=title, font=("Arial",9), fg=Theme.get("SUBTEXT"),
                 bg=Theme.get("CARD")).pack(anchor="w")
        lbl = tk.Label(card, text=value, font=("Arial",15,"bold"),
                       fg=color, bg=Theme.get("CARD"))
        lbl.pack(anchor="w")
        return lbl

    def _refresh_cards(self):
        if not hasattr(self,"_bal_lbl"): return
        self._bal_lbl.config(text=fmt(self.balance))
        self._inc_lbl.config(text=fmt(self.total_income))
        self._exp_lbl.config(text=fmt(self.total_expense))
        pct = (self.balance/self.total_income*100) if self.total_income>0 else 0
        self._sav_lbl.config(text=f"{pct:.1f}%")

    # ─── Analytics ────────────────────────────────────────────────────

    def _build_analytics_page(self, frame):
        bar = tk.Frame(frame, bg=Theme.get("PANEL"), pady=10)
        bar.pack(fill="x")
        tk.Label(bar, text="📊  Monthly Analytics", font=("Arial",14,"bold"),
                 fg=Theme.get("ACCENT"), bg=Theme.get("PANEL")).pack(side="left",padx=20)

        # ── month arrow nav ──
        nav = tk.Frame(bar, bg=Theme.get("PANEL"))
        nav.pack(side="right", padx=20)
        tk.Button(nav, text="◀", font=("Arial",11), bg=Theme.get("CARD"),
                  fg=Theme.get("TEXT"), relief="flat", cursor="hand2",
                  command=lambda: self._ana_nav(-1)).pack(side="left", padx=2)
        self._ana_month_lbl = tk.Label(nav, text="", font=("Arial",11,"bold"),
                                       fg=Theme.get("TEXT"), bg=Theme.get("PANEL"), width=14)
        self._ana_month_lbl.pack(side="left")
        tk.Button(nav, text="▶", font=("Arial",11), bg=Theme.get("CARD"),
                  fg=Theme.get("TEXT"), relief="flat", cursor="hand2",
                  command=lambda: self._ana_nav(1)).pack(side="left", padx=2)

        now = datetime.now()
        self._ana_year  = now.year
        self._ana_month = now.month

        self._ana_stats_frame = tk.Frame(frame, bg=Theme.get("BG"))
        self._ana_stats_frame.pack(fill="x", padx=16, pady=8)
        self._ana_chart_frame = tk.Frame(frame, bg=Theme.get("BG"))
        self._ana_chart_frame.pack(fill="both", expand=True, padx=16, pady=4)

    def _ana_nav(self, delta):
        self._ana_month += delta
        if self._ana_month > 12: self._ana_month=1;  self._ana_year+=1
        elif self._ana_month < 1: self._ana_month=12; self._ana_year-=1
        self._refresh_analytics()

    def _refresh_analytics(self):
        if not hasattr(self,"_ana_stats_frame"): return
        y, m = self._ana_year, self._ana_month
        ym = f"{y}-{m:02d}"
        if hasattr(self,"_ana_month_lbl"):
            self._ana_month_lbl.config(text=f"{calendar.month_name[m]} {y}")

        month_inc=0.0; month_exp=0.0
        daily_exp = defaultdict(float)
        cat_exp   = defaultdict(float)
        monthly_data = defaultdict(lambda:{"income":0.0,"expense":0.0})

        for tx in self.all_transactions:
            try: d=datetime.strptime(tx["Date"],"%d-%m-%Y")
            except: continue
            kym = d.strftime("%Y-%m")
            monthly_data[kym]["income" if tx["Type"]=="Income" else "expense"] += float(tx["Amount"])
            if kym==ym:
                amt=float(tx["Amount"])
                if tx["Type"]=="Income": month_inc+=amt
                else: month_exp+=amt; cat_exp[tx["Category"]]+=amt

        days_elapsed = min(date.today().day, calendar.monthrange(y,m)[1]) \
                       if (y==date.today().year and m==date.today().month) \
                       else calendar.monthrange(y,m)[1]
        daily_avg  = month_exp/days_elapsed if days_elapsed>0 else 0
        top_cat    = max(cat_exp,key=cat_exp.get) if cat_exp else "N/A"
        savings_pct= (month_inc-month_exp)/month_inc*100 if month_inc>0 else 0

        for w in self._ana_stats_frame.winfo_children(): w.destroy()
        stats = [
            ("Monthly Income",   fmt(month_inc),          Theme.get("GREEN")),
            ("Monthly Expenses", fmt(month_exp),          Theme.get("RED")),
            ("Daily Average",    fmt(daily_avg),          Theme.get("ORANGE")),
            ("Top Category",     top_cat,                 Theme.get("ACCENT")),
            ("Savings %",        f"{savings_pct:.1f}%",   Theme.get("GREEN") if savings_pct>=0 else Theme.get("RED")),
            ("Net Balance",      fmt(month_inc-month_exp),Theme.get("ACCENT")),
        ]
        for title,val,color in stats:
            c=tk.Frame(self._ana_stats_frame,bg=Theme.get("CARD"),padx=14,pady=10)
            c.pack(side="left",expand=True,fill="both",padx=4)
            tk.Label(c,text=title,font=("Arial",8),fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD")).pack(anchor="w")
            tk.Label(c,text=val,font=("Arial",13,"bold"),fg=color,bg=Theme.get("CARD")).pack(anchor="w")

        for w in self._ana_chart_frame.winfo_children(): w.destroy()
        fig,axes=plt.subplots(1,3,figsize=(13,3.8))
        fig.patch.set_facecolor(Theme.get("BG"))
        cc=Theme.get("CHART")

        ax1=axes[0]; ax1.set_facecolor(Theme.get("BG"))
        if cat_exp:
            vals=list(cat_exp.values()); lbls=list(cat_exp.keys())
            wedges,texts,autotexts=ax1.pie(vals,labels=lbls,autopct="%1.1f%%",startangle=90,
                colors=cc[:len(vals)],wedgeprops={"edgecolor":Theme.get("BG"),"linewidth":1.5})
            for t in texts:     t.set_color(Theme.get("TEXT")); t.set_fontsize(7)
            for t in autotexts: t.set_color(Theme.get("TEXT")); t.set_fontsize(7)
        ax1.set_title(f"Category Split\n{ym}",color=Theme.get("TEXT"),fontsize=9,fontweight="bold")

        ax2=axes[1]; ax2.set_facecolor(Theme.get("CARD"))
        trend_months=sorted(monthly_data.keys())[-6:]
        t_inc=[monthly_data[k]["income"]  for k in trend_months]
        t_exp=[monthly_data[k]["expense"] for k in trend_months]
        x=range(len(trend_months))
        ax2.plot(list(x),t_inc,color=Theme.get("GREEN"),marker="o",linewidth=2,label="Income")
        ax2.plot(list(x),t_exp,color=Theme.get("RED"),  marker="o",linewidth=2,label="Expense")
        ax2.set_xticks(list(x))
        ax2.set_xticklabels([k[-5:] for k in trend_months],color=Theme.get("SUBTEXT"),fontsize=7,rotation=30)
        ax2.yaxis.set_tick_params(labelcolor=Theme.get("SUBTEXT"))
        ax2.set_title("6-Month Trend",color=Theme.get("TEXT"),fontsize=9,fontweight="bold")
        ax2.legend(fontsize=7,facecolor=Theme.get("CARD"),labelcolor=Theme.get("TEXT"))
        for sp in ax2.spines.values(): sp.set_color(Theme.get("BORDER"))
        ax2.grid(color=Theme.get("BORDER"),linewidth=0.5)

        ax3=axes[2]; ax3.set_facecolor(Theme.get("CARD"))
        bars=ax3.bar(["Income","Expense","Balance"],
                     [month_inc,month_exp,month_inc-month_exp],
                     color=[Theme.get("GREEN"),Theme.get("RED"),Theme.get("ACCENT")],
                     edgecolor=Theme.get("BG"),linewidth=1.2)
        ax3.set_title("Month Summary",color=Theme.get("TEXT"),fontsize=9,fontweight="bold")
        ax3.tick_params(colors=Theme.get("SUBTEXT"))
        for sp in ax3.spines.values(): sp.set_color(Theme.get("BORDER"))
        ax3.grid(axis="y",color=Theme.get("BORDER"),linewidth=0.5)
        for bar in bars:
            ax3.text(bar.get_x()+bar.get_width()/2,
                     bar.get_height()+max(month_inc,month_exp,1)*0.02,
                     fmt(bar.get_height()),ha="center",va="bottom",
                     color=Theme.get("TEXT"),fontsize=7)

        plt.tight_layout(pad=2)
        canvas=FigureCanvasTkAgg(fig,master=self._ana_chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both",expand=True)
        plt.close(fig)

    # ─── Budget ───────────────────────────────────────────────────────

    def _build_budget_page(self, frame):
        bar=tk.Frame(frame,bg=Theme.get("PANEL"),pady=10); bar.pack(fill="x")
        tk.Label(bar,text="💰  Budget Manager",font=("Arial",14,"bold"),
                 fg=Theme.get("ACCENT"),bg=Theme.get("PANEL")).pack(side="left",padx=20)

        form=tk.Frame(frame,bg=Theme.get("CARD"),padx=20,pady=14); form.pack(fill="x",padx=16,pady=10)
        tk.Label(form,text="Set Monthly Budget Limit",font=("Arial",11,"bold"),
                 fg=Theme.get("TEXT"),bg=Theme.get("CARD")).pack(anchor="w",pady=(0,8))
        row=tk.Frame(form,bg=Theme.get("CARD")); row.pack(fill="x")
        tk.Label(row,text="Category:",font=("Arial",9),fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD")).pack(side="left")
        self._bud_cat=ttk.Combobox(row,values=self.EXPENSE_CATS,state="readonly",width=14)
        self._bud_cat.pack(side="left",padx=(6,14)); self._bud_cat.current(0)
        tk.Label(row,text="Limit (₹):",font=("Arial",9),fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD")).pack(side="left")
        self._bud_amt=tk.Entry(row,bg=Theme.get("BG"),fg=Theme.get("TEXT"),
                               insertbackground=Theme.get("TEXT"),relief="flat",font=("Arial",10),width=10)
        self._bud_amt.pack(side="left",padx=(6,14),ipady=4)
        btn=tk.Button(row,text="Set Budget",font=("Arial",10,"bold"),
                      bg=Theme.get("ACCENT"),fg=Theme.get("BG"),relief="flat",
                      padx=14,pady=5,cursor="hand2",command=self._set_budget)
        btn.pack(side="left")
        Tooltip(btn,"Save budget limit for selected category")

        self._bud_progress_frame=tk.Frame(frame,bg=Theme.get("BG"))
        self._bud_progress_frame.pack(fill="both",expand=True,padx=16,pady=6)

    def _set_budget(self):
        cat=self._bud_cat.get()
        try:
            limit=float(self._bud_amt.get())
            if limit<=0: raise ValueError
        except ValueError:
            messagebox.showerror("Invalid","Enter a positive amount."); return
        self.budgets[cat]=limit; self.cfg["budgets"]=self.budgets; save_config(self.cfg)
        self._refresh_budget_page()
        self._set_status(f"Budget set: {cat} → {fmt(limit)}/month")

    def _refresh_budget_page(self):
        if not hasattr(self,"_bud_progress_frame"): return
        for w in self._bud_progress_frame.winfo_children(): w.destroy()
        now=datetime.now(); ym=now.strftime("%Y-%m")
        cat_spent=defaultdict(float)
        for tx in self.all_transactions:
            try: d=datetime.strptime(tx["Date"],"%d-%m-%Y")
            except: continue
            if d.strftime("%Y-%m")==ym and tx["Type"]=="Expense":
                cat_spent[tx["Category"]]+=float(tx["Amount"])
        if not self.budgets:
            tk.Label(self._bud_progress_frame,text="No budgets set yet. Add one above.",
                     font=("Arial",11),fg=Theme.get("SUBTEXT"),bg=Theme.get("BG")).pack(pady=40); return
        tk.Label(self._bud_progress_frame,text=f"Budget Usage — {now.strftime('%B %Y')}",
                 font=("Arial",11,"bold"),fg=Theme.get("TEXT"),bg=Theme.get("BG")).pack(anchor="w",pady=(4,8))
        for cat,limit in self.budgets.items():
            spent=cat_spent.get(cat,0.0); pct=min(spent/limit,1.0) if limit>0 else 0
            row=tk.Frame(self._bud_progress_frame,bg=Theme.get("CARD"),padx=16,pady=10)
            row.pack(fill="x",pady=3)
            head=tk.Frame(row,bg=Theme.get("CARD")); head.pack(fill="x")
            tk.Label(head,text=cat,font=("Arial",10,"bold"),fg=Theme.get("TEXT"),bg=Theme.get("CARD")).pack(side="left")
            color=Theme.get("GREEN") if pct<0.75 else Theme.get("ORANGE") if pct<1.0 else Theme.get("RED")
            status_txt="✅ OK" if pct<0.75 else "⚠️ Nearing" if pct<1.0 else "🚨 Exceeded!"
            tk.Label(head,text=status_txt,font=("Arial",9,"bold"),fg=color,bg=Theme.get("CARD")).pack(side="right")
            tk.Label(row,text=f"{fmt(spent)} of {fmt(limit)}  ({pct*100:.1f}%)",
                     font=("Arial",9),fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD")).pack(anchor="w",pady=(2,4))
            cvs=tk.Canvas(row,height=12,bg=Theme.get("BG"),highlightthickness=0)
            cvs.pack(fill="x"); cvs.update_idletasks()
            w=cvs.winfo_width() or 400
            cvs.create_rectangle(0,0,w,12,fill=Theme.get("BORDER"),outline="")
            cvs.create_rectangle(0,0,int(w*pct),12,fill=color,outline="")

    # ─── Calendar ─────────────────────────────────────────────────────

    def _build_calendar_page(self, frame):
        bar=tk.Frame(frame,bg=Theme.get("PANEL"),pady=10); bar.pack(fill="x")
        tk.Label(bar,text="📅  Calendar View",font=("Arial",14,"bold"),
                 fg=Theme.get("ACCENT"),bg=Theme.get("PANEL")).pack(side="left",padx=20)
        ctrl=tk.Frame(bar,bg=Theme.get("PANEL")); ctrl.pack(side="right",padx=20)
        tk.Button(ctrl,text="◀",font=("Arial",11),bg=Theme.get("CARD"),
                  fg=Theme.get("TEXT"),relief="flat",cursor="hand2",
                  command=lambda: self._cal_nav(-1)).pack(side="left",padx=2)
        self._cal_label=tk.Label(ctrl,text="",font=("Arial",11,"bold"),
                                  fg=Theme.get("TEXT"),bg=Theme.get("PANEL"),width=14)
        self._cal_label.pack(side="left")
        tk.Button(ctrl,text="▶",font=("Arial",11),bg=Theme.get("CARD"),
                  fg=Theme.get("TEXT"),relief="flat",cursor="hand2",
                  command=lambda: self._cal_nav(1)).pack(side="left",padx=2)
        self._cal_year=date.today().year; self._cal_month=date.today().month
        self._cal_grid_frame=tk.Frame(frame,bg=Theme.get("BG"))
        self._cal_grid_frame.pack(fill="both",expand=True,padx=16,pady=8)
        self._cal_detail_frame=tk.Frame(frame,bg=Theme.get("CARD"),pady=10)
        self._cal_detail_frame.pack(fill="x",padx=16,pady=(0,8))

    def _cal_nav(self,delta):
        self._cal_month+=delta
        if self._cal_month>12: self._cal_month=1; self._cal_year+=1
        elif self._cal_month<1: self._cal_month=12; self._cal_year-=1
        self._refresh_calendar()

    def _refresh_calendar(self):
        if not hasattr(self,"_cal_grid_frame"): return
        y,m=self._cal_year,self._cal_month
        self._cal_label.config(text=f"{calendar.month_name[m]} {y}")
        for w in self._cal_grid_frame.winfo_children(): w.destroy()
        for w in self._cal_detail_frame.winfo_children(): w.destroy()

        daily=defaultdict(lambda:{"inc":0.0,"exp":0.0})
        for tx in self.all_transactions:
            try: d=datetime.strptime(tx["Date"],"%d-%m-%Y")
            except: continue
            if d.year==y and d.month==m:
                amt=float(tx["Amount"])
                if tx["Type"]=="Income": daily[d.day]["inc"]+=amt
                else: daily[d.day]["exp"]+=amt
        max_exp=max((v["exp"] for v in daily.values()),default=0)

        CELL_W=12  # uniform column weight

        # Day-of-week headers
        for i,dn in enumerate(["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]):
            tk.Label(self._cal_grid_frame,text=dn,font=("Arial",9,"bold"),
                     fg=Theme.get("ACCENT"),bg=Theme.get("BG"),
                     anchor="center").grid(row=0,column=i,padx=2,pady=2,sticky="ew")

        cal_m=calendar.monthcalendar(y,m)
        for r,week in enumerate(cal_m,start=1):
            for c,day in enumerate(week):
                if day==0:
                    # Empty cell — same fixed size as active cells
                    tk.Frame(self._cal_grid_frame,bg=Theme.get("BG"),
                             width=110,height=70).grid(row=r,column=c,padx=2,pady=2,sticky="nsew")
                    continue

                exp=daily[day]["exp"]; inc=daily[day]["inc"]
                is_today = (y == date.today().year and m == date.today().month and day == date.today().day)

                # ── Calendar cell colouring ──
                # Income-only  → green background
                # Expense-only → intensity-based red/orange
                # Both present → neutral card, show both in colour text
                # Empty        → neutral card
                if inc > 0 and exp == 0:
                    bg_c     = "#1B6B3A"   # deep green — readable on all themes
                    txt_c    = "#FFFFFF"
                    border_c = Theme.get("ACCENT") if is_today else "#26A65B"
                elif exp > 0 and inc == 0:
                    intensity = exp / max_exp if max_exp > 0 else 0
                    if intensity > 0.7:
                        bg_c = "#8B0000"; txt_c = "#FFFFFF"
                        border_c = Theme.get("ACCENT") if is_today else "#CC0000"
                    elif intensity > 0.3:
                        bg_c = "#CC5500"; txt_c = "#FFFFFF"
                        border_c = Theme.get("ACCENT") if is_today else "#FF7700"
                    else:
                        bg_c = "#5A1A1A"; txt_c = "#FFAAAA"
                        border_c = Theme.get("ACCENT") if is_today else bg_c
                else:
                    # both or empty — neutral
                    bg_c     = Theme.get("CARD")
                    txt_c    = Theme.get("TEXT")
                    border_c = Theme.get("ACCENT") if is_today else Theme.get("BORDER")

                # Outer border frame
                cell=tk.Frame(self._cal_grid_frame,bg=border_c,padx=1,pady=1,
                              width=110,height=70)
                cell.grid(row=r,column=c,padx=2,pady=2,sticky="nsew")
                cell.grid_propagate(False)

                inner=tk.Frame(cell,bg=bg_c,padx=6,pady=4)
                inner.place(relwidth=1,relheight=1)

                # Day number
                tk.Label(inner,text=str(day),font=("Arial",10,"bold"),
                         fg=Theme.get("ACCENT") if is_today else txt_c,
                         bg=bg_c,anchor="nw").pack(anchor="nw")

                # Show ONLY totals (no per-transaction detail in the cell)
                if exp>0:
                    exp_fg = txt_c if bg_c != Theme.get("CARD") else "#EF4444"
                    tk.Label(inner,text=f"−{fmt(exp)}",font=("Arial",8),
                             fg=exp_fg, bg=bg_c).pack(anchor="w")
                if inc>0:
                    inc_fg = txt_c if bg_c != Theme.get("CARD") else "#22C55E"
                    tk.Label(inner,text=f"+{fmt(inc)}",font=("Arial",8),
                             fg=inc_fg, bg=bg_c).pack(anchor="w")

                # Bind click AFTER all child labels exist
                def _bind_cell(c=cell, i=inner, d=day):
                    click_cb = lambda e, day=d: self._cal_day_click(day)
                    c.bind("<Button-1>", click_cb)
                    i.bind("<Button-1>", click_cb)
                    for child in i.winfo_children():
                        child.bind("<Button-1>", click_cb)
                _bind_cell()

        for c in range(7):
            self._cal_grid_frame.columnconfigure(c,weight=1)

        tk.Label(self._cal_detail_frame,text="Click a date to see its transactions",
                 font=("Arial",9),fg=Theme.get("SUBTEXT"),
                 bg=Theme.get("CARD"),padx=14).pack(anchor="w")

    def _cal_day_click(self,day):
        for w in self._cal_detail_frame.winfo_children(): w.destroy()
        date_str=f"{day:02d}-{self._cal_month:02d}-{self._cal_year:04d}"
        txs=[tx for tx in self.all_transactions if tx["Date"]==date_str]
        tk.Label(self._cal_detail_frame,text=f"Transactions on {date_str}:",
                 font=("Arial",10,"bold"),fg=Theme.get("TEXT"),
                 bg=Theme.get("CARD"),padx=14).pack(anchor="w",pady=(6,4))
        if not txs:
            tk.Label(self._cal_detail_frame,text="  No transactions.",
                     font=("Arial",9),fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD")).pack(anchor="w"); return
        for tx in txs:
            color=Theme.get("GREEN") if tx["Type"]=="Income" else Theme.get("RED")
            row=tk.Frame(self._cal_detail_frame,bg=Theme.get("CARD")); row.pack(fill="x",padx=14,pady=1)
            tk.Label(row,text=f"{'▲' if tx['Type']=='Income' else '▼'} {tx['Category']}",
                     font=("Arial",9),fg=color,bg=Theme.get("CARD")).pack(side="left")
            tk.Label(row,text=fmt(float(tx["Amount"])),font=("Arial",9,"bold"),
                     fg=color,bg=Theme.get("CARD")).pack(side="left",padx=10)
            if tx.get("Notes"):
                tk.Label(row,text=f"— {tx['Notes']}",font=("Arial",8),
                         fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD")).pack(side="left")

    # ─── Goals ────────────────────────────────────────────────────────

    def _build_goals_page(self, frame):
        bar=tk.Frame(frame,bg=Theme.get("PANEL"),pady=10); bar.pack(fill="x")
        tk.Label(bar,text="🎯  Financial Goals",font=("Arial",14,"bold"),
                 fg=Theme.get("ACCENT"),bg=Theme.get("PANEL")).pack(side="left",padx=20)

        form=tk.Frame(frame,bg=Theme.get("CARD"),padx=20,pady=14)
        form.pack(fill="x",padx=16,pady=10)
        tk.Label(form,text="Create Savings Goal",font=("Arial",11,"bold"),
                 fg=Theme.get("TEXT"),bg=Theme.get("CARD")).pack(anchor="w",pady=(0,8))
        row=tk.Frame(form,bg=Theme.get("CARD")); row.pack(fill="x")
        for lbl,var_name,width in [
            ("Goal Name:","_goal_name_var",16),
            ("Target (₹):","_goal_target_var",10),
            ("Saved (₹):","_goal_saved_var",10),
        ]:
            tk.Label(row,text=lbl,font=("Arial",9),fg=Theme.get("SUBTEXT"),
                     bg=Theme.get("CARD")).pack(side="left",padx=(0,4))
            v=tk.StringVar(); setattr(self,var_name,v)
            tk.Entry(row,textvariable=v,width=width,bg=Theme.get("BG"),fg=Theme.get("TEXT"),
                     insertbackground=Theme.get("TEXT"),relief="flat",
                     font=("Arial",10)).pack(side="left",padx=(0,12),ipady=4)
        tk.Button(row,text="Add Goal",font=("Arial",10,"bold"),
                  bg=Theme.get("ACCENT"),fg=Theme.get("BG"),relief="flat",
                  padx=14,pady=5,cursor="hand2",command=self._add_goal).pack(side="left")

        self._goals_list_frame=tk.Frame(frame,bg=Theme.get("BG"))
        self._goals_list_frame.pack(fill="both",expand=True,padx=16,pady=6)

    def _add_goal(self):
        name=self._goal_name_var.get().strip()
        try:
            target=float(self._goal_target_var.get())
            saved=float(self._goal_saved_var.get() or "0")
        except ValueError:
            messagebox.showerror("Invalid","Enter valid numbers."); return
        if not name: messagebox.showerror("Invalid","Enter a goal name."); return
        self.goals.append({"name":name,"target":target,"saved":saved})
        self.cfg["goals"]=self.goals; save_config(self.cfg)
        self._refresh_goals_page()

    def _refresh_goals_page(self):
        if not hasattr(self,"_goals_list_frame"): return
        for w in self._goals_list_frame.winfo_children(): w.destroy()
        if not self.goals:
            tk.Label(self._goals_list_frame,text="No goals yet. Create one above!",
                     font=("Arial",11),fg=Theme.get("SUBTEXT"),bg=Theme.get("BG")).pack(pady=40); return

        for i,goal in enumerate(self.goals):
            target=float(goal["target"]); saved=float(goal["saved"])
            pct=min(saved/target,1.0) if target>0 else 0
            remaining=max(target-saved,0)
            color=Theme.get("GREEN") if pct>=1.0 else Theme.get("ACCENT")

            card=tk.Frame(self._goals_list_frame,bg=Theme.get("CARD"),padx=16,pady=12)
            card.pack(fill="x",pady=4)
            head=tk.Frame(card,bg=Theme.get("CARD")); head.pack(fill="x")
            tk.Label(head,text=f"🎯 {goal['name']}",font=("Arial",11,"bold"),
                     fg=color,bg=Theme.get("CARD")).pack(side="left")

            if pct>=1.0:
                tk.Label(head,text="✅ ACHIEVED!",font=("Arial",9,"bold"),
                         fg=Theme.get("GREEN"),bg=Theme.get("CARD")).pack(side="right")
            else:
                # Delete button
                tk.Button(head,text="✕",font=("Arial",8),bg=Theme.get("CARD"),
                          fg=Theme.get("RED"),relief="flat",cursor="hand2",
                          command=lambda idx=i: self._delete_goal(idx)).pack(side="right",padx=2)
                # Edit button — opens inline edit
                tk.Button(head,text="✏ Edit",font=("Arial",8),bg=Theme.get("CARD"),
                          fg=Theme.get("ACCENT"),relief="flat",cursor="hand2",
                          command=lambda idx=i,g=goal: self._edit_goal_popup(idx,g)).pack(side="right",padx=4)

            tk.Label(card,text=f"{fmt(saved)} saved of {fmt(target)}  ({pct*100:.1f}%)",
                     font=("Arial",9),fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD")).pack(anchor="w",pady=(4,4))

            cvs=tk.Canvas(card,height=14,bg=Theme.get("BG"),highlightthickness=0)
            cvs.pack(fill="x"); cvs.update_idletasks()
            w=cvs.winfo_width() or 500
            cvs.create_rectangle(0,0,w,14,fill=Theme.get("BORDER"),outline="")
            cvs.create_rectangle(0,0,int(w*pct),14,fill=color,outline="")

            if pct<1.0:
                tk.Label(card,text=f"Remaining: {fmt(remaining)}",
                         font=("Arial",8),fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD")).pack(anchor="w",pady=(4,0))

    def _edit_goal_popup(self, idx, goal):
        popup=tk.Toplevel(self.root)
        popup.title("Edit Goal"); popup.configure(bg=Theme.get("BG"))
        popup.resizable(False,False); popup.grab_set()
        center_window(popup,320,260,relative_to=self.root)

        tk.Label(popup,text="Edit Goal",font=("Arial",13,"bold"),
                 fg=Theme.get("ACCENT"),bg=Theme.get("BG")).pack(pady=(18,12))
        form=tk.Frame(popup,bg=Theme.get("BG")); form.pack(padx=26,fill="x")

        def lbl(t):
            tk.Label(form,text=t,font=("Arial",9),fg=Theme.get("SUBTEXT"),
                     bg=Theme.get("BG")).pack(anchor="w",pady=(8,2))
        def ent(prefill=""):
            e=tk.Entry(form,bg=Theme.get("CARD"),fg=Theme.get("TEXT"),
                       insertbackground=Theme.get("TEXT"),relief="flat",font=("Arial",11))
            e.pack(fill="x",ipady=6); e.insert(0,prefill); return e

        lbl("Goal Name")
        name_e=ent(goal["name"])
        lbl("Target (₹)")
        target_e=ent(str(goal["target"]))
        lbl("Amount Saved (₹)")
        saved_e=ent(str(goal["saved"]))

        btn_row=tk.Frame(popup,bg=Theme.get("BG")); btn_row.pack(pady=14,fill="x",padx=26)
        def do_save():
            name=name_e.get().strip()
            try:
                target=float(target_e.get()); saved=float(saved_e.get())
                if target<=0: raise ValueError
            except ValueError:
                messagebox.showerror("Invalid","Enter valid positive numbers.",parent=popup); return
            if not name: messagebox.showerror("Invalid","Enter a name.",parent=popup); return
            self.goals[idx]={"name":name,"target":target,"saved":saved}
            self.cfg["goals"]=self.goals; save_config(self.cfg)
            popup.destroy(); self._refresh_goals_page()
            self._set_status(f"Goal '{name}' updated.")

        tk.Button(btn_row,text="Cancel",font=("Arial",10),bg=Theme.get("CARD"),
                  fg=Theme.get("TEXT"),relief="flat",padx=14,pady=9,
                  cursor="hand2",command=popup.destroy).pack(side="left",expand=True,fill="x",padx=(0,6))
        tk.Button(btn_row,text="Save  ✓",font=("Arial",10,"bold"),bg=Theme.get("ACCENT"),
                  fg=Theme.get("BG"),relief="flat",padx=14,pady=9,
                  cursor="hand2",command=do_save).pack(side="left",expand=True,fill="x")
        popup.bind("<Return>",lambda _: do_save())

    def _delete_goal(self, idx):
        if messagebox.askyesno("Delete Goal","Remove this goal?"):
            self.goals.pop(idx); self.cfg["goals"]=self.goals; save_config(self.cfg)
            self._refresh_goals_page()

    # ─── Recurring ────────────────────────────────────────────────────

    def _build_recurring_page(self, frame):
        bar=tk.Frame(frame,bg=Theme.get("PANEL"),pady=10); bar.pack(fill="x")
        tk.Label(bar,text="🔄  Recurring Transactions",font=("Arial",14,"bold"),
                 fg=Theme.get("ACCENT"),bg=Theme.get("PANEL")).pack(side="left",padx=20)

        form=tk.Frame(frame,bg=Theme.get("CARD"),padx=20,pady=14); form.pack(fill="x",padx=16,pady=10)
        tk.Label(form,text="Add Recurring Entry",font=("Arial",11,"bold"),
                 fg=Theme.get("TEXT"),bg=Theme.get("CARD")).pack(anchor="w",pady=(0,8))
        row=tk.Frame(form,bg=Theme.get("CARD")); row.pack(fill="x")

        tk.Label(row,text="Name:",font=("Arial",9),fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD")).pack(side="left")
        self._rec_name=tk.Entry(row,width=14,bg=Theme.get("BG"),fg=Theme.get("TEXT"),
                                 insertbackground=Theme.get("TEXT"),relief="flat",font=("Arial",10))
        self._rec_name.pack(side="left",padx=(4,10),ipady=4)
        tk.Label(row,text="Type:",font=("Arial",9),fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD")).pack(side="left")
        self._rec_type=ttk.Combobox(row,values=["Income","Expense"],state="readonly",width=9)
        self._rec_type.current(1); self._rec_type.pack(side="left",padx=(4,10))
        tk.Label(row,text="Category:",font=("Arial",9),fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD")).pack(side="left")
        self._rec_cat=ttk.Combobox(row,values=self.EXPENSE_CATS,state="readonly",width=12)
        self._rec_cat.current(0); self._rec_cat.pack(side="left",padx=(4,10))
        tk.Label(row,text="Amount:",font=("Arial",9),fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD")).pack(side="left")
        self._rec_amt=tk.Entry(row,width=8,bg=Theme.get("BG"),fg=Theme.get("TEXT"),
                                insertbackground=Theme.get("TEXT"),relief="flat",font=("Arial",10))
        self._rec_amt.pack(side="left",padx=(4,10),ipady=4)
        tk.Label(row,text="Freq:",font=("Arial",9),fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD")).pack(side="left")
        self._rec_freq=ttk.Combobox(row,values=["Daily","Weekly","Monthly"],state="readonly",width=9)
        self._rec_freq.current(2); self._rec_freq.pack(side="left",padx=(4,10))
        tk.Button(row,text="Add",font=("Arial",10,"bold"),bg=Theme.get("ACCENT"),fg=Theme.get("BG"),
                  relief="flat",padx=12,pady=5,cursor="hand2",command=self._add_recurring).pack(side="left")

        self._rec_list_frame=tk.Frame(frame,bg=Theme.get("BG"))
        self._rec_list_frame.pack(fill="both",expand=True,padx=16,pady=6)
        self._refresh_recurring_page()

    def _add_recurring(self):
        name=self._rec_name.get().strip()
        try:
            amt=float(self._rec_amt.get())
            if amt<=0: raise ValueError
        except ValueError:
            messagebox.showerror("Invalid","Enter a valid positive amount."); return
        if not name: messagebox.showerror("Invalid","Enter a name."); return
        self.recurring.append({"name":name,"type":self._rec_type.get(),
                               "category":self._rec_cat.get(),"amount":amt,
                               "frequency":self._rec_freq.get(),"last_run":""})
        self.cfg["recurring"]=self.recurring; save_config(self.cfg)
        self._refresh_recurring_page()

    def _refresh_recurring_page(self):
        if not hasattr(self,"_rec_list_frame"): return
        for w in self._rec_list_frame.winfo_children(): w.destroy()
        if not self.recurring:
            tk.Label(self._rec_list_frame,text="No recurring transactions. Add one above.",
                     font=("Arial",11),fg=Theme.get("SUBTEXT"),bg=Theme.get("BG")).pack(pady=30); return
        cols=("Name","Type","Category","Amount","Frequency","Last Run")
        tree=ttk.Treeview(self._rec_list_frame,columns=cols,show="headings",height=12)
        widths={"Name":140,"Type":80,"Category":120,"Amount":100,"Frequency":100,"Last Run":110}
        for col in cols:
            tree.heading(col,text=col); tree.column(col,width=widths[col],anchor="center")
        for r in self.recurring:
            tree.insert("",tk.END,values=(r["name"],r["type"],r["category"],
                fmt(r["amount"]),r["frequency"],r.get("last_run","Never")))
        tree.pack(fill="both",expand=True)
        tk.Button(self._rec_list_frame,text="🗑  Remove Selected",
                  font=("Arial",9),bg=Theme.get("RED"),fg=Theme.get("TEXT"),
                  relief="flat",padx=10,pady=4,cursor="hand2",
                  command=lambda: self._remove_recurring(tree)).pack(pady=6)

    def _remove_recurring(self, tree):
        sel=tree.selection()
        if not sel: return
        for i in sorted([tree.index(it) for it in sel],reverse=True): self.recurring.pop(i)
        self.cfg["recurring"]=self.recurring; save_config(self.cfg)
        self._refresh_recurring_page()

    def _process_recurring(self):
        today_str=date.today().strftime("%d-%m-%Y"); today=date.today(); changed=False
        for r in self.recurring:
            last=r.get("last_run",""); should_run=False
            if not last: should_run=True
            else:
                try:
                    ld=datetime.strptime(last,"%d-%m-%Y").date(); freq=r.get("frequency","Monthly")
                    if   freq=="Daily"   and (today-ld).days>=1:  should_run=True
                    elif freq=="Weekly"  and (today-ld).days>=7:  should_run=True
                    elif freq=="Monthly" and (today.year>ld.year or today.month>ld.month): should_run=True
                except: should_run=True
            if should_run:
                with open(self.FILE,"a",newline="") as fh:
                    csv.writer(fh).writerow([today_str,r["type"],r["category"],r["amount"],f"Auto: {r['name']}"])
                r["last_run"]=today_str; changed=True
        if changed: self.cfg["recurring"]=self.recurring; save_config(self.cfg)

    # ─── Export ───────────────────────────────────────────────────────

    def _build_export_page(self, frame):
        bar=tk.Frame(frame,bg=Theme.get("PANEL"),pady=10); bar.pack(fill="x")
        tk.Label(bar,text="📤  Export Reports",font=("Arial",14,"bold"),
                 fg=Theme.get("ACCENT"),bg=Theme.get("PANEL")).pack(side="left",padx=20)

        content=tk.Frame(frame,bg=Theme.get("BG")); content.pack(fill="both",expand=True,padx=30,pady=20)

        options=[
            ("📊  Export to Excel (.xlsx)",self._export_excel,
             "Full transaction history with formatted table and summary sheet.",HAS_XLSX),
            ("📄  Export to PDF",self._export_pdf,
             "Formatted PDF report with transaction table and financial summary.",HAS_PDF),
            ("💾  Backup CSV",self._export_csv,
             "Save a timestamped backup of your raw transaction data.",True),
        ]
        for label,cmd,desc,available in options:
            card=tk.Frame(content,bg=Theme.get("CARD"),padx=20,pady=16); card.pack(fill="x",pady=6)
            head=tk.Frame(card,bg=Theme.get("CARD")); head.pack(fill="x")
            tk.Label(head,text=label,font=("Arial",12,"bold"),
                     fg=Theme.get("ACCENT") if available else Theme.get("SUBTEXT"),
                     bg=Theme.get("CARD")).pack(side="left")
            if available:
                b=tk.Button(head,text="Export",font=("Arial",10,"bold"),
                            bg=Theme.get("ACCENT"),fg=Theme.get("BG"),relief="flat",
                            padx=14,pady=5,cursor="hand2",command=cmd)
                b.pack(side="right"); Tooltip(b,"Export now")
            else:
                pip_cmd = "pip install openpyxl" if "Excel" in label else "pip install reportlab"
                tk.Label(head,text=f"Run: {pip_cmd}",
                         font=("Consolas",8),fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD")).pack(side="right")
            tk.Label(card,text=desc,font=("Arial",9),fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD")).pack(anchor="w",pady=(6,0))

    def _export_excel(self):
        if not HAS_XLSX: messagebox.showerror("Not Available","openpyxl is not installed.\nRun: pip install openpyxl"); return
        filename=f"trackify_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="Transactions"
        hf=Font(bold=True,color="000000",size=11)
        hfill=PatternFill("solid",fgColor="00CCCC")
        ca=Alignment(horizontal="center")
        for c,h in enumerate(["Date","Type","Category","Amount","Notes"],1):
            cell=ws.cell(row=1,column=c,value=h); cell.font=hf; cell.fill=hfill; cell.alignment=ca
        for r,tx in enumerate(self.all_transactions,2):
            rf=PatternFill("solid",fgColor="D4EDDA" if tx["Type"]=="Income" else "F8D7DA")
            for c,key in enumerate(["Date","Type","Category","Amount","Notes"],1):
                cell=ws.cell(row=r,column=c,value=float(tx["Amount"]) if key=="Amount" else tx.get(key,""))
                cell.fill=rf
                if key=="Amount": cell.number_format="₹#,##0.00"
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=18
        ws2=wb.create_sheet("Summary")
        ws2.cell(1,1,"Trackify Financial Summary").font=Font(bold=True,size=14)
        ws2.cell(2,1,f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}").font=Font(color="888888")
        for r,(lbl,val) in enumerate([("Total Income",self.total_income),
                                       ("Total Expenses",self.total_expense),
                                       ("Balance",self.balance)],4):
            ws2.cell(r,1,lbl).font=Font(bold=True); ws2.cell(r,2,val)
        wb.save(filename)
        messagebox.showinfo("Exported",f"Excel saved:\n{filename}")
        self._set_status(f"Exported to {filename}")

    def _export_pdf(self):
        if not HAS_PDF: messagebox.showerror("Not Available","reportlab not installed.\nRun: pip install reportlab"); return
        filename=f"trackify_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        doc=SimpleDocTemplate(filename,pagesize=A4,leftMargin=40,rightMargin=40,topMargin=40,bottomMargin=40)
        styles=getSampleStyleSheet(); elements=[]
        elements.append(Paragraph("Trackify — Financial Report",styles["Title"]))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}",styles["Normal"]))
        elements.append(Spacer(1,0.3*inch))
        elements.append(Paragraph("Financial Summary",styles["Heading2"]))
        st=Table([["Metric","Value"],
                  ["Total Income",fmt(self.total_income)],
                  ["Total Expenses",fmt(self.total_expense)],
                  ["Net Balance",fmt(self.balance)],
                  ["Savings %",f"{(self.balance/self.total_income*100) if self.total_income>0 else 0:.1f}%"]],
                 colWidths=[3*inch,2*inch])
        st.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),rl_colors.HexColor("#00CCCC")),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[rl_colors.white,rl_colors.HexColor("#F5F5F5")]),
            ("GRID",(0,0),(-1,-1),0.5,rl_colors.grey),
        ]))
        elements.append(st); elements.append(Spacer(1,0.3*inch))
        elements.append(Paragraph("All Transactions",styles["Heading2"]))
        tx_data=[["Date","Type","Category","Amount","Notes"]]
        for tx in self.all_transactions:
            tx_data.append([tx["Date"],tx["Type"],tx["Category"],fmt(float(tx["Amount"])),tx.get("Notes","")[:30]])
        if len(tx_data)>1:
            tt=Table(tx_data,colWidths=[1.2*inch,0.9*inch,1.4*inch,1.2*inch,2.3*inch])
            row_colors=[("BACKGROUND",(0,i),(-1,i),
                         rl_colors.HexColor("#E8F5E9") if tx["Type"]=="Income" else rl_colors.HexColor("#FFEBEE"))
                        for i,tx in enumerate(self.all_transactions,1)]
            tt.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,0),rl_colors.HexColor("#00CCCC")),
                ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                ("FONTSIZE",(0,0),(-1,-1),8),("ALIGN",(0,0),(-1,-1),"CENTER"),
                ("GRID",(0,0),(-1,-1),0.3,rl_colors.grey),*row_colors]))
            elements.append(tt)
        doc.build(elements)
        messagebox.showinfo("Exported",f"PDF saved:\n{filename}")
        self._set_status(f"Exported to {filename}")

    def _export_csv(self):
        filename=f"trackify_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        shutil.copy2(self.FILE,filename)
        messagebox.showinfo("Backed Up",f"CSV backup saved:\n{filename}")
        self._set_status(f"CSV backup: {filename}")

    # ─── Settings ─────────────────────────────────────────────────────

    def _build_settings_page(self, frame):
        bar=tk.Frame(frame,bg=Theme.get("PANEL"),pady=10); bar.pack(fill="x")
        tk.Label(bar,text="⚙️  Settings",font=("Arial",14,"bold"),
                 fg=Theme.get("ACCENT"),bg=Theme.get("PANEL")).pack(side="left",padx=20)

        # Scrollable settings area
        outer=tk.Frame(frame,bg=Theme.get("BG")); outer.pack(fill="both",expand=True)
        cvs=tk.Canvas(outer,bg=Theme.get("BG"),highlightthickness=0)
        vsb=ttk.Scrollbar(outer,orient="vertical",command=cvs.yview)
        cvs.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right",fill="y"); cvs.pack(side="left",fill="both",expand=True)
        content=tk.Frame(cvs,bg=Theme.get("BG"))
        cw=cvs.create_window((0,0),window=content,anchor="nw")
        content.bind("<Configure>",lambda e: cvs.configure(scrollregion=cvs.bbox("all")))
        cvs.bind("<Configure>",lambda e: cvs.itemconfig(cw,width=e.width))

        def _settings_scroll(e):
            cvs.yview_scroll(-1*(e.delta//120), "units")
        cvs.bind("<MouseWheel>", _settings_scroll)
        content.bind("<MouseWheel>", _settings_scroll)

        def _bind_settings_scroll(widget):
            widget.bind("<MouseWheel>", _settings_scroll)
            for child in widget.winfo_children():
                _bind_settings_scroll(child)

        def section(title, emoji=""):
            f=tk.Frame(content,bg=Theme.get("CARD"),padx=20,pady=14)
            f.pack(fill="x",padx=20,pady=6)
            tk.Label(f,text=f"{emoji}  {title}",font=("Arial",11,"bold"),
                     fg=Theme.get("TEXT"),bg=Theme.get("CARD")).pack(anchor="w",pady=(0,10))
            return f

        def row(parent):
            r=tk.Frame(parent,bg=Theme.get("CARD")); r.pack(fill="x",pady=3); return r

        # ── 1. Currency & Display ──
        s1=section("Currency & Display","💱")
        r1=row(s1)
        tk.Label(r1,text="Currency Symbol:",font=("Arial",9),fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD"),width=20,anchor="w").pack(side="left")
        self._cfg_currency=tk.StringVar(value=self.cfg.get("currency","₹"))
        currency_cb=ttk.Combobox(r1,textvariable=self._cfg_currency,
                    values=["₹","$","€","£","¥","₩","₺","₦","₫","฿"],
                    state="readonly",width=5,font=("Arial",11))
        currency_cb.pack(side="left",padx=8,ipady=4)
        tk.Label(r1,text="(e.g. ₹  $  €  £)",font=("Arial",8),fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD")).pack(side="left")

        r2=row(s1)
        tk.Label(r2,text="Date Format:",font=("Arial",9),fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD"),width=20,anchor="w").pack(side="left")
        self._cfg_datefmt=tk.StringVar(value=self.cfg.get("date_format","dd-mm-yyyy"))
        dfcb=ttk.Combobox(r2,textvariable=self._cfg_datefmt,
                          values=["dd-mm-yyyy","mm-dd-yyyy","yyyy-mm-dd"],state="readonly",width=14)
        dfcb.pack(side="left",padx=8)

        # ── 2. Dashboard defaults ──
        s2=section("Dashboard Defaults","🏠")
        r3=row(s2)
        tk.Label(r3,text="Default Filter:",font=("Arial",9),fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD"),width=20,anchor="w").pack(side="left")
        self._cfg_def_filter=tk.StringVar(value=self.cfg.get("default_filter","All"))
        ttk.Combobox(r3,textvariable=self._cfg_def_filter,
                     values=["All","Income","Expense"],state="readonly",width=10).pack(side="left",padx=8)

        r4=row(s2)
        tk.Label(r4,text="Start on current month:",font=("Arial",9),fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD"),width=20,anchor="w").pack(side="left")
        self._cfg_cur_month=tk.BooleanVar(value=self.cfg.get("start_current_month",True))
        tk.Checkbutton(r4,variable=self._cfg_cur_month,bg=Theme.get("CARD"),
                       fg=Theme.get("TEXT"),selectcolor=Theme.get("BG"),
                       activebackground=Theme.get("CARD")).pack(side="left",padx=8)

        # ── 3. Notifications / Alerts ──
        s3=section("Budget Alerts","🔔")
        r5=row(s3)
        tk.Label(r5,text="Warn at % of budget:",font=("Arial",9),fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD"),width=20,anchor="w").pack(side="left")
        self._cfg_warn_pct=tk.StringVar(value=str(self.cfg.get("budget_warn_pct",80)))
        tk.Entry(r5,textvariable=self._cfg_warn_pct,width=5,bg=Theme.get("BG"),
                 fg=Theme.get("TEXT"),insertbackground=Theme.get("TEXT"),relief="flat",font=("Arial",11)).pack(side="left",padx=8,ipady=4)
        tk.Label(r5,text="%",font=("Arial",9),fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD")).pack(side="left")

        r6=row(s3)
        tk.Label(r6,text="Enable budget alerts:",font=("Arial",9),fg=Theme.get("SUBTEXT"),bg=Theme.get("CARD"),width=20,anchor="w").pack(side="left")
        self._cfg_alerts=tk.BooleanVar(value=self.cfg.get("budget_alerts",True))
        tk.Checkbutton(r6,variable=self._cfg_alerts,bg=Theme.get("CARD"),
                       fg=Theme.get("TEXT"),selectcolor=Theme.get("BG"),
                       activebackground=Theme.get("CARD")).pack(side="left",padx=8)

        # ── 4. Data Management ──
        s4=section("Data Management","🗄️")
        dm_row=tk.Frame(s4,bg=Theme.get("CARD")); dm_row.pack(fill="x",pady=4)
        for btn_txt, btn_cmd, btn_tip in [
            ("📦  Backup CSV now", self._export_csv, "Save timestamped CSV backup"),
            ("🗑  Clear ALL data", self._clear_all_data, "⚠ Permanently delete all transactions"),
            ("📂  Open data file", self._open_data_file, "Open transactions.csv in your default app"),
        ]:
            is_danger = "Clear" in btn_txt
            fg_col = Theme.get("RED") if is_danger else Theme.get("TEXT")
            bg_col = Theme.get("BG")
            b=tk.Button(dm_row,text=btn_txt,font=("Arial",9,"bold"),
                        bg=bg_col,fg=fg_col,
                        relief="solid",bd=1,padx=14,pady=8,cursor="hand2",command=btn_cmd,
                        activebackground=Theme.get("RED") if is_danger else Theme.get("ACCENT"),
                        activeforeground=Theme.get("BG"),highlightthickness=1,
                        highlightbackground=Theme.get("RED") if is_danger else Theme.get("BORDER"))
            b.pack(side="left",padx=(0,10))
            b.bind("<Enter>", lambda e, btn=b, danger=is_danger: btn.config(
                bg=Theme.get("RED") if danger else Theme.get("ACCENT"),
                fg=Theme.get("BG")))
            b.bind("<Leave>", lambda e, btn=b, fg=fg_col: btn.config(bg=bg_col, fg=fg))
            Tooltip(b,btn_tip)

        # ── 5. Keyboard shortcuts reference ──
        s5=section("Keyboard Shortcuts","⌨️")
        for key,desc in [
            ("Ctrl+N","Add new transaction"),
            ("Ctrl+F","Focus search bar (hint inside box)"),
            ("Del","Delete via ✕ button on each row"),
            ("Ctrl+Z","Undo last delete"),
            ("Ctrl+E","Go to Export page"),
            ("F5","Refresh dashboard"),
            ("Double-click row","Edit transaction"),
        ]:
            kr=tk.Frame(s5,bg=Theme.get("CARD")); kr.pack(fill="x",pady=2)
            tk.Label(kr,text=key,font=("Consolas",9,"bold"),fg=Theme.get("ACCENT"),
                     bg=Theme.get("BORDER"),padx=8,pady=2,width=22,anchor="w").pack(side="left")
            tk.Label(kr,text=desc,font=("Arial",9),fg=Theme.get("SUBTEXT"),
                     bg=Theme.get("CARD"),padx=12).pack(side="left")

        # ── 6. About ──
        s6=section("About Trackify","ℹ️")
        for k,v in [("Version","3.2"),("Author","Trackify Project"),
                    ("Data file","transactions.csv"),("Config file","trackify_config.json")]:
            ar=tk.Frame(s6,bg=Theme.get("CARD")); ar.pack(fill="x",pady=1)
            tk.Label(ar,text=k+":",font=("Arial",9,"bold"),fg=Theme.get("TEXT"),
                     bg=Theme.get("CARD"),width=16,anchor="w").pack(side="left")
            tk.Label(ar,text=v,font=("Consolas",9),fg=Theme.get("ACCENT"),
                     bg=Theme.get("CARD")).pack(side="left",padx=8)

        # ── Save Settings button ──
        save_row=tk.Frame(content,bg=Theme.get("BG")); save_row.pack(fill="x",padx=20,pady=12)
        tk.Button(save_row,text="💾  Save Settings",font=("Arial",11,"bold"),
                  bg=Theme.get("ACCENT"),fg=Theme.get("BG"),relief="flat",
                  padx=20,pady=10,cursor="hand2",command=self._save_settings).pack(side="left")

        # Bind mousewheel to everything inside settings after full build
        content.after(100, lambda: _bind_settings_scroll(content))

    def _save_settings(self):
        try:
            warn_pct=int(self._cfg_warn_pct.get())
            if not (1<=warn_pct<=100): raise ValueError
        except ValueError:
            messagebox.showerror("Invalid","Budget warn % must be 1–100."); return
        self.cfg["currency"]           = self._cfg_currency.get() or "₹"
        self.cfg["date_format"]        = self._cfg_datefmt.get()
        self.cfg["default_filter"]     = self._cfg_def_filter.get()
        self.cfg["start_current_month"]= self._cfg_cur_month.get()
        self.cfg["budget_warn_pct"]    = warn_pct
        self.cfg["budget_alerts"]      = self._cfg_alerts.get()
        save_config(self.cfg)
        messagebox.showinfo("Saved","Settings saved.\nSome changes take effect on next launch.")

    def _clear_all_data(self):
        if not messagebox.askyesno("⚠ Clear ALL Data",
            "This will permanently delete ALL transactions.\nThis cannot be undone!\n\nAre you sure?",
            parent=self.root): return
        with open(self.FILE,"w",newline="") as fh:
            csv.writer(fh).writerow(self.COLS)
        self._recalculate(); self._load_transactions(reload=True)
        messagebox.showinfo("Cleared","All transaction data has been deleted.")
        self._set_status("All data cleared.")

    def _open_data_file(self):
        import subprocess, sys
        try:
            if sys.platform=="win32":  os.startfile(self.FILE)
            elif sys.platform=="darwin": subprocess.call(["open",self.FILE])
            else: subprocess.call(["xdg-open",self.FILE])
        except Exception as e:
            messagebox.showerror("Error",f"Could not open file:\n{e}")

    # ─── Sort (hidden table) ──────────────────────────────────────────

    def _sort_by(self, col):
        asc=not self._sort_state.get(col,True); self._sort_state[col]=asc
        data=[(self.table.set(c,col),c) for c in self.table.get_children("")]
        try:   data.sort(key=lambda x: float(x[0].replace("₹","").replace(",","")),reverse=not asc)
        except:data.sort(reverse=not asc)
        for i,(_,child) in enumerate(data): self.table.move(child,"",i)

    # ─── Delete / Undo ────────────────────────────────────────────────

    def _delete_selected(self):
        """Delete rows selected in the hidden treeview."""
        sel=self.table.selection()
        if not sel:
            messagebox.showinfo("Nothing Selected","Select one or more rows to delete."); return
        if not messagebox.askyesno("Confirm Delete",f"Permanently delete {len(sel)} transaction(s)?",parent=self.root): return
        to_delete=[]
        for item in sel:
            v=self.table.item(item,"values")
            raw_amt=v[3].replace("₹","").replace(",","")
            to_delete.append((v[0],v[1],v[2],raw_amt,v[4]))
            self.table.delete(item)
        kept,deleted_rows=[],[]
        with open(self.FILE,"r") as fh:
            for row in csv.DictReader(fh):
                key=(row["Date"],row["Type"],row["Category"],row["Amount"],row.get("Notes",""))
                matched=False
                for i,dk in enumerate(to_delete):
                    if key==dk: to_delete.pop(i); deleted_rows.append(row); matched=True; break
                if not matched: kept.append(row)
        with open(self.FILE,"w",newline="") as fh:
            w=csv.DictWriter(fh,fieldnames=list(self.COLS)); w.writeheader(); w.writerows(kept)
        self._undo_buffer=deleted_rows
        self._recalculate(); self._refresh_dash_table()
        self._set_status(f"Deleted {len(sel)} transaction(s).")
        self._show_undo_bar(len(sel))

    def _delete_tx_direct(self, tx):
        """Delete a single transaction dict from the day-view rows."""
        if not messagebox.askyesno("Delete","Delete this transaction?",parent=self.root): return
        kept=[]
        matched=False
        with open(self.FILE,"r") as fh:
            for row in csv.DictReader(fh):
                if (not matched and row["Date"]==tx["Date"] and row["Type"]==tx["Type"]
                        and row["Category"]==tx["Category"] and row["Amount"]==tx["Amount"]):
                    matched=True
                else:
                    kept.append(row)
        with open(self.FILE,"w",newline="") as fh:
            w=csv.DictWriter(fh,fieldnames=list(self.COLS)); w.writeheader(); w.writerows(kept)
        self._recalculate(); self._refresh_dash_table()
        self._set_status("Transaction deleted.")

    def _show_undo_bar(self, count):
        if not hasattr(self,"_undo_bar"): return
        self._undo_lbl.config(text=f"  Deleted {count} transaction(s).")
        self._undo_bar.pack(fill="x")
        if self._undo_job: self.root.after_cancel(self._undo_job)
        self._undo_job=self.root.after(6000,self._hide_undo_bar)

    def _hide_undo_bar(self):
        if hasattr(self,"_undo_bar"): self._undo_bar.pack_forget()
        self._undo_buffer=None

    def _undo_delete(self):
        if not self._undo_buffer: self._set_status("Nothing to undo."); return
        with open(self.FILE,"a",newline="") as fh:
            w=csv.DictWriter(fh,fieldnames=list(self.COLS))
            for row in self._undo_buffer: w.writerow(row)
        self._undo_buffer=None; self._hide_undo_bar()
        self._recalculate(); self._load_transactions(reload=True)
        self._set_status("Undo successful.")

    # ─── Edit ─────────────────────────────────────────────────────────

    def _edit_selected(self):
        sel=self.table.selection()
        if not sel: messagebox.showinfo("Nothing Selected","Select a transaction to edit."); return
        if len(sel)>1: messagebox.showinfo("One at a Time","Select a single transaction to edit."); return
        vals=self.table.item(sel[0],"values")
        tx={"Date":vals[0],"Type":vals[1],"Category":vals[2],
            "Amount":vals[3].replace("₹","").replace(",",""),"Notes":vals[4]}
        self._edit_tx_direct(tx)

    def _edit_tx_direct(self, tx):
        popup=tk.Toplevel(self.root)
        popup.title("Edit Transaction"); popup.configure(bg=Theme.get("BG"))
        popup.resizable(False,False); popup.grab_set()
        center_window(popup,340,430,relative_to=self.root)

        tk.Label(popup,text="Edit Transaction",font=("Arial",13,"bold"),
                 fg=Theme.get("ACCENT"),bg=Theme.get("BG")).pack(pady=(18,12))
        form=tk.Frame(popup,bg=Theme.get("BG")); form.pack(padx=26,fill="x")

        def lbl(t):
            tk.Label(form,text=t,font=("Arial",9),fg=Theme.get("SUBTEXT"),bg=Theme.get("BG")).pack(anchor="w",pady=(8,2))
        def ent(prefill=""):
            e=tk.Entry(form,bg=Theme.get("CARD"),fg=Theme.get("TEXT"),
                       insertbackground=Theme.get("TEXT"),relief="flat",font=("Arial",11))
            e.pack(fill="x",ipady=6); e.insert(0,prefill); return e

        lbl("Date")
        date_e=DateEntry(form,width=28,background=Theme.get("CARD"),
                         foreground=Theme.get("TEXT"),borderwidth=0,date_pattern="dd-mm-yyyy")
        date_e.pack(fill="x",ipady=4)
        try: date_e.set_date(datetime.strptime(tx["Date"],"%d-%m-%Y"))
        except: date_e.set_date(datetime.now())

        lbl("Type")
        type_box=ttk.Combobox(form,values=["Income","Expense"],state="readonly")
        type_box.pack(fill="x",ipady=4); type_box.set(tx["Type"])

        lbl("Category")
        cats=self.INCOME_CATS if tx["Type"]=="Income" else self.EXPENSE_CATS
        cat_box=ttk.Combobox(form,values=cats,state="readonly")
        cat_box.pack(fill="x",ipady=4); cat_box.set(tx["Category"])

        def on_type_change(_=None):
            cat_box["values"]=self.INCOME_CATS if type_box.get()=="Income" else self.EXPENSE_CATS
            cat_box.current(0)
        type_box.bind("<<ComboboxSelected>>",on_type_change)

        lbl("Amount (₹)"); amt_e=ent(tx["Amount"])
        lbl("Notes");      notes_e=ent(tx.get("Notes",""))

        btn_row=tk.Frame(popup,bg=Theme.get("BG")); btn_row.pack(pady=16,fill="x",padx=26)

        def do_save():
            try:
                new_amt=float(amt_e.get())
                if new_amt<=0: raise ValueError
            except ValueError:
                messagebox.showerror("Invalid","Enter a positive amount.",parent=popup); return
            old_key=(tx["Date"],tx["Type"],tx["Category"],tx["Amount"],tx.get("Notes",""))
            kept=[]; replaced=False
            with open(self.FILE,"r") as fh:
                for row in csv.DictReader(fh):
                    key=(row["Date"],row["Type"],row["Category"],row["Amount"],row.get("Notes",""))
                    if not replaced and key==old_key:
                        kept.append({"Date":date_e.get(),"Type":type_box.get(),
                                     "Category":cat_box.get(),"Amount":new_amt,
                                     "Notes":notes_e.get().strip()})
                        replaced=True
                    else: kept.append(row)
            with open(self.FILE,"w",newline="") as fh:
                w=csv.DictWriter(fh,fieldnames=list(self.COLS)); w.writeheader(); w.writerows(kept)
            self._recalculate(); self._load_transactions(reload=True)
            popup.destroy(); self._set_status("Transaction updated.")

        tk.Button(btn_row,text="Cancel",font=("Arial",10),bg=Theme.get("CARD"),
                  fg=Theme.get("TEXT"),relief="flat",padx=14,pady=9,
                  cursor="hand2",command=popup.destroy).pack(side="left",expand=True,fill="x",padx=(0,6))
        tk.Button(btn_row,text="Save  ✓",font=("Arial",10,"bold"),bg=Theme.get("ACCENT"),
                  fg=Theme.get("BG"),relief="flat",padx=14,pady=9,
                  cursor="hand2",command=do_save).pack(side="left",expand=True,fill="x")
        popup.bind("<Return>",lambda _: do_save())
        popup.bind("<Escape>",lambda _: popup.destroy())

    def _open_add_window(self):
        popup=tk.Toplevel(self.root)
        popup.title("New Transaction"); popup.configure(bg=Theme.get("BG"))
        popup.resizable(False,False); popup.grab_set()
        center_window(popup,340,440,relative_to=self.root)

        tk.Label(popup,text="Add Transaction",font=("Arial",13,"bold"),
                 fg=Theme.get("ACCENT"),bg=Theme.get("BG")).pack(pady=(18,12))
        form=tk.Frame(popup,bg=Theme.get("BG")); form.pack(padx=26,fill="x")

        def lbl(t):
            tk.Label(form,text=t,font=("Arial",9),fg=Theme.get("SUBTEXT"),bg=Theme.get("BG")).pack(anchor="w",pady=(8,2))
        def ent():
            e=tk.Entry(form,bg=Theme.get("CARD"),fg=Theme.get("TEXT"),
                       insertbackground=Theme.get("TEXT"),relief="flat",font=("Arial",11))
            e.pack(fill="x",ipady=6); return e

        lbl("Date")
        date_entry=DateEntry(form,width=28,background=Theme.get("CARD"),
                             foreground=Theme.get("TEXT"),borderwidth=0,date_pattern="dd-mm-yyyy")
        date_entry.pack(fill="x",ipady=4); date_entry.set_date(datetime.now())

        lbl("Type")
        type_box=ttk.Combobox(form,values=["Income","Expense"],state="readonly")
        type_box.pack(fill="x",ipady=4); type_box.current(1)

        lbl("Category")
        cat_box=ttk.Combobox(form,values=self.EXPENSE_CATS,state="readonly")
        cat_box.pack(fill="x",ipady=4); cat_box.current(0)

        def on_type_change(_=None):
            cat_box["values"]=self.INCOME_CATS if type_box.get()=="Income" else self.EXPENSE_CATS
            cat_box.current(0)
        type_box.bind("<<ComboboxSelected>>",on_type_change)

        lbl("Amount (₹)"); amt_entry=ent(); amt_entry.focus()
        lbl("Notes (optional)"); notes_entry=ent()

        btn_row=tk.Frame(popup,bg=Theme.get("BG")); btn_row.pack(pady=18,fill="x",padx=26)

        def do_add():
            try:
                amount=float(amt_entry.get())
                if amount<=0: raise ValueError
            except ValueError:
                messagebox.showerror("Invalid Amount","Please enter a positive number.",parent=popup); return
            d=date_entry.get(); t_type=type_box.get()
            category=cat_box.get(); notes=notes_entry.get().strip()
            tx={"Date":d,"Type":t_type,"Category":category,"Amount":str(amount),"Notes":notes}
            self.all_transactions.append(tx)
            if t_type=="Income":
                self.balance+=amount; self.total_income+=amount
            else:
                self.balance-=amount; self.total_expense+=amount
                self.expenses[category]=self.expenses.get(category,0)+amount
            self._refresh_cards()
            # also insert into hidden table for delete/edit compatibility
            tag="income" if t_type=="Income" else "expense"
            self.table.insert("",tk.END,values=(d,t_type,category,fmt(amount),notes),tags=(tag,))
            with open(self.FILE,"a",newline="") as fh:
                csv.writer(fh).writerow([d,t_type,category,amount,notes])
            popup.destroy()
            self._set_status(f"Added {t_type}: {fmt(amount)} — {category}")
            self._refresh_dash_table()
            self._check_budget_alert(category,t_type)

        tk.Button(btn_row,text="Cancel",font=("Arial",10),bg=Theme.get("CARD"),
                  fg=Theme.get("TEXT"),relief="flat",padx=14,pady=9,
                  cursor="hand2",command=popup.destroy).pack(side="left",expand=True,fill="x",padx=(0,6))
        tk.Button(btn_row,text="Add  →",font=("Arial",10,"bold"),bg=Theme.get("ACCENT"),
                  fg=Theme.get("BG"),relief="flat",padx=14,pady=9,
                  cursor="hand2",command=do_add,activebackground="#66FFFF").pack(side="left",expand=True,fill="x")
        popup.bind("<Return>",lambda _: do_add())
        popup.bind("<Escape>",lambda _: popup.destroy())

    def _check_budget_alert(self, category, t_type):
        if t_type!="Expense" or category not in self.budgets: return
        limit=self.budgets[category]; spent=self.expenses.get(category,0); pct=spent/limit if limit>0 else 0
        if   pct>=1.0: messagebox.showwarning("🚨 Budget Exceeded!",f"You've exceeded your {category} budget!\nSpent: {fmt(spent)} / Limit: {fmt(limit)}")
        elif pct>=0.8: messagebox.showwarning("⚠️ Budget Warning",f"You're at {pct*100:.0f}% of your {category} budget.\nSpent: {fmt(spent)} / Limit: {fmt(limit)}")

    # ─── Chart ────────────────────────────────────────────────────────

    def _show_chart(self):
        if not self.expenses:
            messagebox.showinfo("No Expenses","No expense data yet."); return

        labels = list(self.expenses.keys())
        values = list(self.expenses.values())
        colors = Theme.get("CHART")[:len(labels)]

        win = tk.Toplevel(self.root)
        win.title("Trackify — Spending Charts")
        win.configure(bg=Theme.get("BG"))
        center_window(win, 1020, 520, relative_to=self.root)
        win.resizable(True, True)

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(11, 4.8))
        fig.patch.set_facecolor(Theme.get("BG"))

        # ── Create canvas FIRST so _draw_pie can reference it ──
        canvas = FigureCanvasTkAgg(fig, master=win)
        canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=6)

        summary_txt = (f"  Total Expenses: {fmt(self.total_expense)}   |   "
                       f"Total Income: {fmt(self.total_income)}   |   "
                       f"Balance: {fmt(self.balance)}")
        tk.Label(win, text=summary_txt, font=("Arial",9),
                 fg=Theme.get("SUBTEXT"), bg=Theme.get("BG")).pack(pady=(0,8))

        # ── Smooth animated explode state ──
        _target_explode = [0.0] * len(values)
        _cur_explode    = [0.0] * len(values)
        _anim_job       = [None]
        _wedges         = [None]

        def _draw_pie():
            ax1.clear()
            ax1.set_facecolor(Theme.get("BG"))
            wedges, texts, autotexts = ax1.pie(
                values, labels=labels, autopct="%1.1f%%",
                explode=_cur_explode, startangle=90, colors=colors,
                wedgeprops={"edgecolor": Theme.get("BG"), "linewidth": 2},
                labeldistance=1.12)
            ax1.set_title("Category Distribution",
                          color=Theme.get("TEXT"), fontsize=10, fontweight="bold", pad=12)
            for t in texts:
                t.set_color(Theme.get("TEXT")); t.set_fontsize(8); t.set_weight("bold")
            for t in autotexts:
                t.set_color(Theme.get("TEXT")); t.set_fontsize(8)
            _wedges[0] = wedges
            canvas.draw_idle()

        def _animate_explode():
            changed = False
            for i in range(len(_cur_explode)):
                diff = _target_explode[i] - _cur_explode[i]
                if abs(diff) > 0.002:
                    _cur_explode[i] += diff * 0.28
                    changed = True
                else:
                    _cur_explode[i] = _target_explode[i]
            _draw_pie()
            if changed:
                _anim_job[0] = win.after(16, _animate_explode)
            else:
                _anim_job[0] = None

        def _on_pie_hover(event):
            if event.inaxes != ax1:
                if any(t > 0 for t in _target_explode):
                    for i in range(len(_target_explode)): _target_explode[i] = 0.0
                    if not _anim_job[0]: _animate_explode()
                return
            if _wedges[0] is None: return
            for i, w in enumerate(_wedges[0]):
                if w.contains_point([event.x, event.y]):
                    new = [0.0] * len(values); new[i] = 0.18
                    if new != _target_explode:
                        for j in range(len(_target_explode)): _target_explode[j] = new[j]
                        if not _anim_job[0]: _animate_explode()
                    return
            if any(t > 0 for t in _target_explode):
                for i in range(len(_target_explode)): _target_explode[i] = 0.0
                if not _anim_job[0]: _animate_explode()

        # ── Bar chart ──
        ax2.set_facecolor(Theme.get("CARD"))
        bar_rects = ax2.bar(labels, values, color=colors,
                            edgecolor=Theme.get("BG"), linewidth=1.5, zorder=2)
        ax2.grid(axis="y", color=Theme.get("BORDER"), linewidth=0.8, zorder=1)
        ax2.set_title("By Category", color=Theme.get("TEXT"), fontsize=11, fontweight="bold")
        ax2.tick_params(colors=Theme.get("TEXT"))
        for sp in ["top","right"]: ax2.spines[sp].set_visible(False)
        for sp in ["bottom","left"]: ax2.spines[sp].set_color(Theme.get("BORDER"))
        for l in ax2.get_xticklabels(): l.set_color(Theme.get("TEXT")); l.set_fontsize(8)
        max_v = max(values)
        for bar, val in zip(bar_rects, values):
            ax2.text(bar.get_x()+bar.get_width()/2, bar.get_height()+max_v*0.012,
                     fmt(val), ha="center", va="bottom",
                     color=Theme.get("TEXT"), fontsize=8, fontweight="bold")

        def _on_bar_hover(event):
            if event.inaxes != ax2: return
            changed = False
            for bar in bar_rects:
                hit = bar.contains(event)[0]
                new_lw = 3.0 if hit else 1.5
                new_ec = "white" if hit else Theme.get("BG")
                if bar.get_linewidth() != new_lw:
                    bar.set_linewidth(new_lw); bar.set_edgecolor(new_ec); changed = True
            if changed: fig.canvas.draw_idle()

        fig.canvas.mpl_connect("motion_notify_event", _on_pie_hover)
        fig.canvas.mpl_connect("motion_notify_event", _on_bar_hover)
        fig.suptitle("Trackify — Spending Overview",
                     color=Theme.get("ACCENT"), fontsize=13, fontweight="bold")
        plt.tight_layout(pad=2.2)

        # Draw initial pie AFTER canvas is ready
        _draw_pie()
        canvas.draw()

        def _on_win_close():
            if _anim_job[0]: win.after_cancel(_anim_job[0])
            plt.close(fig); win.destroy()
        win.protocol("WM_DELETE_WINDOW", _on_win_close)

    # ─── Data ─────────────────────────────────────────────────────────

    def _recalculate(self):
        self.balance=self.total_income=self.total_expense=0.0; self.expenses={}; self.all_transactions=[]
        with open(self.FILE,"r") as fh:
            for row in csv.DictReader(fh):
                amt=float(row["Amount"]); self.all_transactions.append(dict(row))
                if row["Type"]=="Income": self.balance+=amt; self.total_income+=amt
                else:
                    self.balance-=amt; self.total_expense+=amt
                    self.expenses[row["Category"]]=self.expenses.get(row["Category"],0)+amt
        self._refresh_cards()

    def _load_transactions(self, reload=False):
        if reload:
            for row in self.table.get_children(): self.table.delete(row)
            self._recalculate()
            for tx in self.all_transactions:
                tag="income" if tx["Type"]=="Income" else "expense"
                self.table.insert("",tk.END,values=(tx["Date"],tx["Type"],tx["Category"],
                    fmt(float(tx["Amount"])),tx.get("Notes","")),tags=(tag,))
            self._refresh_dash_table(); return
        with open(self.FILE,"r") as fh:
            for row in csv.DictReader(fh):
                amt=float(row["Amount"]); t=row["Type"]; cat=row["Category"]
                self.all_transactions.append(dict(row))
                if t=="Income": self.balance+=amt; self.total_income+=amt; tag="income"
                else:
                    self.balance-=amt; self.total_expense+=amt
                    self.expenses[cat]=self.expenses.get(cat,0)+amt; tag="expense"
                self.table.insert("",tk.END,values=(row["Date"],t,cat,
                    fmt(amt),row.get("Notes","")),tags=(tag,))
        self._refresh_cards()
        self._refresh_dash_table()

    # ─── Misc ─────────────────────────────────────────────────────────

    def _set_status(self, msg):
        if hasattr(self,"_status"): self._status.set(f"  {msg}")
        self.root.after(4000, lambda: self._status.set("  Ready") if hasattr(self,"_status") else None)

    @staticmethod
    def _hover_btn(btn, normal, hover):
        btn.bind("<Enter>", lambda _: btn.config(bg=hover))
        btn.bind("<Leave>", lambda _: btn.config(bg=normal))

    def _on_close(self):
        self.cfg["geometry"]=self.root.geometry().split("+")[0]
        save_config(self.cfg); self.root.destroy()


# ─────────────────────────────────────────────
#  Entry point
# ─────────────────────────────────────────────

if __name__=="__main__":
    root=tk.Tk()
    IntroPage(root)
    root.mainloop()